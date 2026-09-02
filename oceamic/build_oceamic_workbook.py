# -*- coding: utf-8 -*-
"""OCEAMIC - Systeme Qualite / Stock / Sous-traitance.
Builder: genere le classeur Excel (version production + version demo/tests)."""
import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.comments import Comment

# ----------------------------------------------------------------------------
# Capacites (nombre de lignes pre-formatees)
# ----------------------------------------------------------------------------
N_LOTS = 200          # LOTS         lignes 5..204
N_OPS  = 300          # OPERATIONS   lignes 5..304
N_ST   = 150          # SOUS_TRAITANCE lignes 5..154
N_RES  = 300          # ST_RESULTATS lignes 5..304
N_QC   = 250          # QUALITE      lignes 5..254
N_EMPL = 40           # PARAM emplacements lignes 6..45
N_LIST = 40           # PARAM autres listes
N_COLS_EMPL = 12      # colonnes emplacement dans la matrice STOCK (B..M)
HDR = 4               # ligne d'en-tete des sheets de saisie
FIRST = 5             # premiere ligne de donnees

# blocs du grand livre MOUVEMENTS (lignes 5..1054)
MV_B1 = 5                       # 5..304    sorties issues de OPERATIONS
MV_B2 = MV_B1 + N_OPS           # 305..604  entrees issues de OPERATIONS
MV_B3 = MV_B2 + N_OPS           # 605..754  sorties issues de SOUS_TRAITANCE
MV_B4 = MV_B3 + N_ST            # 755..1054 entrees issues de ST_RESULTATS
MV_END = MV_B4 + N_RES - 1      # 1054

# ----------------------------------------------------------------------------
# Charte graphique
# ----------------------------------------------------------------------------
FONT = "Arial"
NAVY      = "0F3D5C"   # bandeau principal
TEAL      = "1B7A8C"   # en-tetes de tableaux
TEAL_SOFT = "D7E7EB"   # bandeau de section
INK       = "1F2933"   # texte principal
MUTED     = "5A6B76"   # texte secondaire
INPUT_BG  = "FFF9E6"   # cellules a saisir
CALC_BG   = "EEF2F5"   # cellules calculees
KPI_BG    = "F4F8F9"
WHITE     = "FFFFFF"
LINE      = "C3CED6"

G_OK   = ("E4F3E7", "14713A")
G_INFO = ("E4EEF9", "1B4F8F")
G_WARN = ("FFF2D6", "8A5A00")
G_BAD  = ("FBE3E3", "A31515")
G_GREY = ("EDEFF1", "5A6B76")

thin = Side(style="thin", color=LINE)
BOX  = Border(left=thin, right=thin, top=thin, bottom=thin)
BOT  = Border(bottom=thin)

NUM_KG   = '#,##0.00;-#,##0.00;'          # zero -> vide
NUM_KG0  = '#,##0.00;-#,##0.00;"-"'
NUM_PCT  = '0.0%'
NUM_DATE = 'dd/mm/yyyy'
NUM_INT  = '#,##0;-#,##0;'


def cell(ws, ref, value=None, bold=False, size=10, color=INK, fill=None,
         align=None, wrap=False, fmt=None, border=None, italic=False, indent=0):
    c = ws[ref]
    if value is not None:
        c.value = value
    c.font = Font(name=FONT, size=size, bold=bold, color=color, italic=italic)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap,
                            indent=indent)
    if fmt:
        c.number_format = fmt
    if border:
        c.border = border
    return c


def title_bar(ws, text, subtitle, last_col="R"):
    ws.merge_cells("A1:%s1" % last_col)
    cell(ws, "A1", text, bold=True, size=15, color=WHITE, fill=NAVY,
         align="left", indent=1)
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:%s2" % last_col)
    cell(ws, "A2", subtitle, size=9, color=MUTED, align="left", indent=1)
    ws.row_dimensions[2].height = 18
    ws.sheet_view.showGridLines = False


def section(ws, row, text, last_col, size=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row,
                   end_column=last_col)
    cell(ws, "A%d" % row, text, bold=True, size=size, color=NAVY,
         fill=TEAL_SOFT, align="left", indent=1)
    ws.row_dimensions[row].height = 20


def headers(ws, row, cols, widths):
    """cols: list of (letter, label, kind) ; kind in {'in','calc'}"""
    for (letter, label, kind) in cols:
        c = cell(ws, "%s%d" % (letter, row), label, bold=True, size=9,
                 color=WHITE, fill=TEAL if kind == "in" else MUTED,
                 align="center", wrap=True, border=BOX)
    ws.row_dimensions[row].height = 30
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w


def body_style(ws, cols, r0, r1, fmts=None):
    fmts = fmts or {}
    for (letter, label, kind) in cols:
        bg = INPUT_BG if kind == "in" else CALC_BG
        col = MUTED if kind == "calc" else INK
        it = kind == "calc"
        for r in range(r0, r1 + 1):
            cell(ws, "%s%d" % (letter, r), None, size=9, color=col, fill=bg,
                 border=BOX, italic=it, fmt=fmts.get(letter))


def fill_col(ws, letter, r0, r1, formula_tpl):
    """formula_tpl uses {r} placeholder."""
    for r in range(r0, r1 + 1):
        ws["%s%d" % (letter, r)] = formula_tpl.format(r=r)


def cf_status(ws, rng):
    rules = [(G_OK, ["OK"]), (G_INFO, ["EN COURS"]), (G_WARN, ["A JUSTIFIER"]),
             (G_GREY, ["DONNEES MANQUANTES"]),
             (G_BAD, ["STOCK INSUFFISANT"]), (G_BAD, ["EMPLACEMENT INCOHERENT"]),
             (G_BAD, ["BLOQUE"]), (G_BAD, ["INCOHERENCE"]),
             (G_WARN, ["LOT EN DOUBLE"])]
    for (bg, fg), vals in rules:
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="equal", formula=['"%s"' % vals[0]],
            fill=PatternFill("solid", fgColor=bg),
            font=Font(name=FONT, size=9, bold=True, color=fg)))


def dv(ws, rng, formula1, kind="list", strict=False, prompt=None, op=None):
    d = DataValidation(type=kind, formula1=formula1, operator=op,
                       allow_blank=True, showErrorMessage=strict,
                       showDropDown=False)
    if strict:
        d.errorTitle = "Valeur non autorisee"
        d.error = "Choisissez une valeur dans la liste deroulante."
    if prompt:
        d.promptTitle = "Aide"
        d.prompt = prompt
        d.showInputMessage = True
    ws.add_data_validation(d)
    d.add(rng)
    return d
# ============================================================================
# PARAM
# ============================================================================
LISTS = [
    ("F", "PRODUIT / ESPECE", "L_PRODUIT",
     ["Sardine", "Maquereau", "Bonite", "Anchois", "Chinchard"]),
    ("H", "PRODUCTEUR / FOURNISSEUR", "L_FOURN",
     ["Fournisseur local", "Producteur Dakhla", "Producteur Boujdour",
      "Autre"]),
    ("J", "ORIGINE", "L_ORIGINE",
     ["Dakhla", "Boujdour", "Laayoune", "Mauritanie", "tan tan"]),
    ("L", "ETAT MATIERE", "L_ETAT",
     ["ENTIER", "HG", "HGT", "FILET", "IQF", "AUTRE"]),
    ("N", "CONSERVATION", "L_CONSERV", ["FRAIS", "CONGELE"]),
    ("P", "QUALITE  (liste d'aide - la saisie libre est acceptee, "
          "ex. A=20 B=40 C=40)", "L_QUALITE",
     ["A", "B", "C", "A=20 B=40 C=40"]),
    ("R", "MOULE / CALIBRE  (liste d'aide - saisie libre acceptee, "
          "ex. 20/24=14 26/30=55 38/50=31)", "L_MOULE",
     ["20/24", "26/30", "30/40", "38/50",
      "20/24=14 26/30=55 38/50=31"]),
    ("T", "SITE DE SOUS-TRAITANCE", "L_STRAIT",
     ["SARMA", "DAMSA", "COFRIGOP", "COFRIGOB", "FOURSEASEN", "ATLANTIC",
      "WILL FISHING", "KJ FISH", "CONGELATION", "USINE SOUS-TRAITANTE"]),
    ("V", "DECISION QUALITE", "L_DECISION",
     ["ACCEPTER", "REFUSER", "A RECONTROLER", "BLOQUER"]),
    ("X", "MOTIF ECART", "L_MOTIF",
     ["PERTE PROCESS", "DECHET / PARURE", "DESHYDRATATION", "CASSE",
      "ERREUR DE PESEE", "ERREUR DE SAISIE", "AUTRE"]),
    ("Z", "TYPE DE CONTROLE", "L_TYPECTRL",
     ["CONTROLE MOUVEMENT", "RECEPTION", "STOCK", "SOUS-TRAITANCE",
      "EXPEDITION"]),
]

EMPLACEMENTS = [
    ("OCEAMIC 2", "INTERNE", "Usine OCEAMIC - seul stock interne"),
    ("SARMA", "EXTERNE", "Entrepot frigorifique"),
    ("DAMSA", "EXTERNE", "Entrepot frigorifique"),
    ("COFRIGOP", "EXTERNE", "Entrepot frigorifique"),
    ("COFRIGOB", "EXTERNE", "Entrepot frigorifique"),
    ("FOURSEASEN", "EXTERNE", "Entrepot frigorifique"),
    ("ATLANTIC", "EXTERNE", "Entrepot frigorifique"),
    ("WILL FISHING", "EXTERNE", "Entrepot frigorifique"),
    ("KJ FISH", "EXTERNE", "Entrepot frigorifique"),
    ("OCEAMIC 1", "EXTERNE", "Site externe"),
]

ENUM_TYPEOP = '"RECEPTION,TRANSFERT,CONSOMMATION,AJUSTEMENT"'
ENUM_SRCST = '"FOURNISSEUR,STOCK EXISTANT"'


def build_param(wb):
    ws = wb.create_sheet("PARAM")
    title_bar(ws, "PARAMETRES  /  LISTES DE REFERENCE",
              "Toutes les listes deroulantes du classeur sont alimentees ici. "
              "Ajoutez vos valeurs a la suite : les menus se mettent a jour "
              "automatiquement.", last_col="AA")
    section(ws, 4, "EMPLACEMENTS DE STOCK  -  la CLASSE ne peut valoir que "
            "INTERNE ou EXTERNE", 27)

    for letter, label in (("B", "EMPLACEMENT"), ("C", "CLASSE"),
                          ("D", "TYPE / COMMENTAIRE")):
        cell(ws, "%s5" % letter, label, bold=True, size=9, color=WHITE,
             fill=TEAL, align="center", border=BOX)
    for i in range(N_EMPL):
        r = 6 + i
        for letter in ("B", "C", "D"):
            cell(ws, "%s%d" % (letter, r), None, size=9, fill=INPUT_BG,
                 border=BOX)
    for i, (e, k, t) in enumerate(EMPLACEMENTS):
        ws["B%d" % (6 + i)] = e
        ws["C%d" % (6 + i)] = k
        ws["D%d" % (6 + i)] = t
    dv(ws, "C6:C%d" % (5 + N_EMPL), '"INTERNE,EXTERNE"', strict=True)
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 2

    for letter, label, name, values in LISTS:
        cell(ws, "%s5" % letter, label, bold=True, size=9, color=WHITE,
             fill=TEAL, align="center", wrap=True, border=BOX)
        for i in range(N_LIST):
            cell(ws, "%s%d" % (letter, 6 + i), None, size=9, fill=INPUT_BG,
                 border=BOX)
        for i, v in enumerate(values):
            ws["%s%d" % (letter, 6 + i)] = v
        ws.column_dimensions[letter].width = 24
        nxt = get_column_letter(ws[letter + "5"].column + 1)
        ws.column_dimensions[nxt].width = 2
    ws.row_dimensions[5].height = 28

    # --- valeurs figees (documentation) ---------------------------------
    section(ws, 48, "VALEURS FIGEES PAR LA LOGIQUE METIER  (non modifiables)",
            27)
    figees = [("TYPE D'OPERATION",
               "RECEPTION / TRANSFERT / CONSOMMATION / AJUSTEMENT"),
              ("CLASSE DE STOCK", "INTERNE / EXTERNE  (aucune autre valeur)"),
              ("SOURCE SOUS-TRAITANCE", "FOURNISSEUR / STOCK EXISTANT"),
              ("STATUTS", "OK / EN COURS / DONNEES MANQUANTES / STOCK "
                          "INSUFFISANT / EMPLACEMENT INCOHERENT / A JUSTIFIER "
                          "/ BLOQUE / LOT EN DOUBLE / INCOHERENCE")]
    for i, (a, b) in enumerate(figees):
        r = 49 + i
        cell(ws, "B%d" % r, a, bold=True, size=9, border=BOX, fill=CALC_BG)
        ws.merge_cells("C%d:J%d" % (r, r))
        cell(ws, "C%d" % r, b, size=9, color=MUTED, border=BOX, fill=CALC_BG,
             align="left", indent=1)

    # --- reglages -------------------------------------------------------
    section(ws, 54, "REGLAGES DU SYSTEME", 27)
    reglages = [
        ("TOLERANCE ECART SOUS-TRAITANCE (kg)", 0, "P_TOL",
         "En dessous de cette valeur, l'ecart n'exige pas de justification."),
        ("SEUIL HISTAMINE (PPM)", 100, "P_HIST",
         "Au-dessus, une alerte est levee dans QUALITE. Jamais bloquant."),
        ("TEMPERATURE MAXI - FRAIS (C)", 4, "P_TFRAIS",
         "Alerte si la temperature relevee depasse ce seuil."),
        ("TEMPERATURE MAXI - CONGELE (C)", -18, "P_TCONG",
         "Alerte si la temperature relevee depasse ce seuil."),
    ]
    for i, (lab, val, name, note) in enumerate(reglages):
        r = 55 + i
        cell(ws, "B%d" % r, lab, bold=True, size=9, border=BOX, fill=CALC_BG)
        cell(ws, "C%d" % r, val, size=10, bold=True, color="1B4F8F",
             fill=INPUT_BG, border=BOX, align="center")
        ws.merge_cells("D%d:J%d" % (r, r))
        cell(ws, "D%d" % r, note, size=8, color=MUTED, border=BOX,
             fill=CALC_BG, align="left", indent=1)
    return ws


def add_names(wb):
    def N(name, ref):
        wb.defined_names.add(DefinedName(name, attr_text=ref))

    def dyn(name, anchor, counta_range):
        N(name, "OFFSET(%s,0,0,MAX(1,COUNTA(%s)),1)" % (anchor, counta_range))

    N("EMPL_KEY", "PARAM!$B$6:$B$%d" % (5 + N_EMPL))
    N("L_PRODUIT_R", "PARAM!$F$6:$F$%d" % (5 + N_LIST))
    N("L_ETAT_R", "PARAM!$L$6:$L$%d" % (5 + N_LIST))
    N("EMPL_CLASSE", "PARAM!$C$6:$C$%d" % (5 + N_EMPL))
    dyn("L_EMPL", "PARAM!$B$6", "PARAM!$B$6:$B$%d" % (5 + N_EMPL))
    for letter, label, name, values in LISTS:
        rng = "PARAM!${0}$6:${0}${1}".format(letter, 5 + N_LIST)
        dyn(name, "PARAM!$%s$6" % letter, rng)
    for i, (lab, val, name, note) in enumerate(
            [("", 0, "P_TOL", ""), ("", 0, "P_HIST", ""),
             ("", 0, "P_TFRAIS", ""), ("", 0, "P_TCONG", "")]):
        N(name, "PARAM!$C$%d" % (55 + i))

    L1, L2 = FIRST, FIRST + N_LOTS - 1
    for name, col in [("LOT_KEY", "A"), ("LOT_EXT", "B"), ("LOT_PROD", "C"),
                      ("LOT_FOURN", "D"), ("LOT_ORIG", "E"),
                      ("LOT_IMMAT", "F"), ("LOT_ETAT", "G"),
                      ("LOT_CONSERV", "H"), ("LOT_QUAL", "I"),
                      ("LOT_MOULE", "J"), ("LOT_DATE", "K"),
                      ("LOT_PARENT", "L"), ("LOT_OBS", "M"),
                      ("LOT_PARENTE", "N"), ("LOT_EXTE", "O"),
                      ("LOT_PRODE", "P"), ("LOT_FOURNE", "Q"),
                      ("LOT_ORIGE", "R"), ("LOT_CONSERVE", "S"),
                      ("LOT_IMMATE", "T"), ("LOT_STOCK", "U"),
                      ("LOT_DECISION", "X"), ("LOT_STATUT", "Y"),
                      ("LOT_RANGEXT", "AA")]:
        N(name, "LOTS!${0}${1}:${0}${2}".format(col, L1, L2))
    dyn("L_LOTS", "LOTS!$A$%d" % L1, "LOTS!$A$%d:$A$%d" % (L1, L2))

    O1, O2 = FIRST, FIRST + N_OPS - 1
    for name, col in [("OPS_ID", "A"), ("OPS_TYPE", "C"), ("OPS_LOT", "D"),
                      ("OPS_STATUT", "Q")]:
        N(name, "OPERATIONS!${0}${1}:${0}${2}".format(col, O1, O2))

    S1, S2 = FIRST, FIRST + N_ST - 1
    for name, col in [("ST_ID", "A"), ("ST_DATE", "B"), ("ST_DATERET", "K"), ("ST_LOT", "E"),
                      ("ST_STRAIT", "C"), ("ST_QTE", "G"), ("ST_STATUT", "X"),
                      ("ST_ENCOURS", "AA"), ("ST_RANG", "AC"),
                      ("ST_ECART", "V"), ("ST_TOTSORTIE", "T"),
                      ("ST_SOURCE", "D"), ("ST_EMPL", "F"), ("ST_NBRES", "U")]:
        N(name, "SOUS_TRAITANCE!${0}${1}:${0}${2}".format(col, S1, S2))
    dyn("L_IDST", "SOUS_TRAITANCE!$A$%d" % S1,
        "SOUS_TRAITANCE!$D$%d:$D$%d" % (S1, S2))

    R1, R2 = FIRST, FIRST + N_RES - 1
    for name, col in [("RES_ID", "A"), ("RES_IDST", "B"), ("RES_DATE", "W"),
                      ("RES_ETAT", "D"), ("RES_QTE", "F"), ("RES_QUAL", "G"),
                      ("RES_MOULE", "H"), ("RES_DEST", "J"),
                      ("RES_LOTSRC", "L"), ("RES_LOTEFF", "N"),
                      ("RES_QTEVAL", "R"), ("RES_STATUT", "T"),
                      ("RES_LOTLIE", "X"), ("RES_RANGLIEN", "Y"),
                      ("RES_RANG", "AA")]:
        N(name, "ST_RESULTATS!${0}${1}:${0}${2}".format(col, R1, R2))

    Q1, Q2 = FIRST, FIRST + N_QC - 1
    for name, col in [("QC_DATE", "B"), ("QC_LOT", "C"), ("QC_EMPL", "D"),
                      ("QC_TEMP", "F"), ("QC_HIST", "G"), ("QC_QUAL", "H"),
                      ("QC_MOULE", "I"), ("QC_DEF", "J"), ("QC_DEC", "K"),
                      ("QC_STATUT", "R"), ("QC_ALERTE", "S"),
                      ("QC_CLEDER", "U"), ("QC_RANG", "W")]:
        N(name, "QUALITE!${0}${1}:${0}${2}".format(col, Q1, Q2))

    for name, col in [("MVT_NO", "A"), ("MVT_SRC", "B"), ("MVT_REF", "C"),
                      ("MVT_DATE", "D"), ("MVT_SENS", "E"), ("MVT_OP", "F"),
                      ("MVT_LOT", "G"), ("MVT_EXT", "H"), ("MVT_EMPL", "I"),
                      ("MVT_CLASSE", "J"), ("MVT_PROD", "K"),
                      ("MVT_ETAT", "L"), ("MVT_BRUT", "M"), ("MVT_NET", "N"),
                      ("MVT_STATUT", "O"), ("MVT_RANG", "P")]:
        N(name, "MOUVEMENTS!${0}${1}:${0}${2}".format(col, MV_B1, MV_END))

    N("R_LOT", "RECHERCHE!$C$5")
# ============================================================================
# LOTS
# ============================================================================
LOTS_COLS = [
    ("A", "LOT INTERNE", "in"), ("B", "LOT EXTERNE", "in"),
    ("C", "ESPECE", "in"), ("D", "PRODUCTEUR / FOURNISSEUR", "in"),
    ("E", "ORIGINE", "in"), ("F", "MATRICULE CAMION", "in"),
    ("G", "ETAT MATIERE", "in"), ("H", "CONSERVATION", "in"),
    ("I", "QUALITE INITIALE", "in"), ("J", "MOULE INITIAL", "in"),
    ("K", "DATE CREATION", "in"), ("L", "LOT PARENT", "in"),
    ("M", "OBSERVATION", "in"),
    ("N", "LOT PARENT EFFECTIF", "calc"), ("O", "LOT EXTERNE EFFECTIF", "calc"),
    ("P", "ESPECE EFFECTIVE", "calc"), ("Q", "PRODUCTEUR EFFECTIF", "calc"),
    ("R", "ORIGINE EFFECTIVE", "calc"),
    ("S", "CONSERVATION EFFECTIVE", "calc"),
    ("T", "MATRICULE EFFECTIF", "calc"), ("U", "STOCK TOTAL (KG)", "calc"),
    ("V", "NB MOUVEMENTS", "calc"), ("W", "NB CONTROLES", "calc"),
    ("X", "DERNIERE DECISION QUALITE", "calc"), ("Y", "STATUT FICHE", "calc"),
    ("Z", "REL. LOT EXTERNE", "calc"), ("AA", "RANG LOT EXTERNE", "calc"),
]
LOTS_W = {"A": 14, "B": 14, "C": 14, "D": 24, "E": 13, "F": 17, "G": 13,
          "H": 15, "I": 20, "J": 26, "K": 12, "L": 12, "M": 30, "N": 16,
          "O": 17, "P": 16, "Q": 22, "R": 15, "S": 18, "T": 17, "U": 14,
          "V": 12, "W": 11, "X": 20, "Y": 20, "Z": 12, "AA": 13}


def build_lots(wb):
    ws = wb.create_sheet("LOTS")
    r0, r1 = FIRST, FIRST + N_LOTS - 1
    title_bar(ws, "LOTS  -  FICHIER D'IDENTITE",
              "Une ligne = un lot. LOT INTERNE est la cle du stock : elle est "
              "obligatoire et ne change plus. Si aucun numero interne n'est "
              "encore attribue, reprenez le LOT EXTERNE du fournisseur. "
              "Ces informations ne se saisissent qu'une seule fois.", "AA")
    section(ws, 3, "Champs jaunes = saisie   |   Champs gris = calcules "
            "automatiquement, ne rien y ecrire", 27)
    headers(ws, HDR, LOTS_COLS, LOTS_W)
    body_style(ws, LOTS_COLS, r0, r1,
               fmts={"K": NUM_DATE, "U": NUM_KG0, "V": NUM_INT, "W": NUM_INT,
                     "Z": NUM_INT})
    ws.freeze_panes = "C%d" % FIRST

    fill_col(ws, "N", r0, r1,
             '=IF($A{r}="","",IF($L{r}<>"",$L{r},'
             'IF(IFERROR(INDEX(RES_LOTSRC,MATCH($A{r},RES_LOTEFF,0)),"")'
             '=$A{r},"",IFERROR(INDEX(RES_LOTSRC,'
             'MATCH($A{r},RES_LOTEFF,0)),""))))')
    for eff, own in (("O", "B"), ("P", "C"), ("Q", "D"), ("R", "E"),
                     ("S", "H"), ("T", "F")):
        fill_col(ws, eff, r0, r1,
                 ('=IF($A{{r}}="","",IF(${o}{{r}}<>"",${o}{{r}},'
                  'IF($N{{r}}="","",IFERROR(INDEX(${o}${a}:${o}${b},'
                  'MATCH($N{{r}},$A${a}:$A${b},0)),""))))'
                  ).format(o=own, a=r0, b=r1))
    fill_col(ws, "U", r0, r1,
             '=IF($A{r}="","",SUMIFS(MVT_NET,MVT_LOT,$A{r}))')
    fill_col(ws, "V", r0, r1,
             '=IF($A{r}="","",COUNTIFS(MVT_LOT,$A{r},MVT_BRUT,"<>0"))')
    fill_col(ws, "W", r0, r1, '=IF($A{r}="","",COUNTIFS(QC_LOT,$A{r}))')
    fill_col(ws, "X", r0, r1,
             '=IF($A{r}="","",IFERROR(INDEX(QC_DEC,'
             'MATCH($A{r},QC_CLEDER,0)),""))')
    fill_col(ws, "Y", r0, r1,
             '=IF($A{r}="","",IF(COUNTIF($A$%d:$A$%d,$A{r})>1,'
             'IF(SUMPRODUCT(($A$%d:$A$%d=$A{r})*(($C$%d:$C$%d<>$C{r})+'
             '($D$%d:$D$%d<>$D{r})))>0,"INCOHERENCE","LOT EN DOUBLE"),"OK"))'
             % (r0, r1, r0, r1, r0, r1, r0, r1))
    fill_col(ws, "Z", r0, r1,
             '=IF($A{r}="",0,IF(AND($O{r}<>"",$A{r}<>R_LOT,'
             '$O{r}=IFERROR(INDEX(LOT_EXTE,MATCH(R_LOT,LOT_KEY,0)),"#")),'
             '1,0))')
    fill_col(ws, "AA", r0, r1,
             '=IF($Z{r}=0,"",COUNTIF($Z$%d:$Z{r},1))' % r0)

    for col, name in [("C", "L_PRODUIT"), ("D", "L_FOURN"), ("E", "L_ORIGINE"),
                      ("G", "L_ETAT"), ("H", "L_CONSERV"), ("I", "L_QUALITE"),
                      ("J", "L_MOULE"), ("L", "L_LOTS")]:
        dv(ws, "{0}{1}:{0}{2}".format(col, r0, r1), name, strict=False)
    cf_status(ws, "Y%d:Y%d" % (r0, r1))
    cf_status(ws, "X%d:X%d" % (r0, r1))
    ws.conditional_formatting.add(
        "A%d:B%d" % (r0, r1),
        FormulaRule(formula=['AND($A%d<>"",$Y%d<>"OK")' % (r0, r0)],
                    fill=PatternFill("solid", fgColor=G_BAD[0])))
    return ws


# ============================================================================
# OPERATIONS
# ============================================================================
OPS_COLS = [
    ("A", "ID OPERATION", "calc"), ("B", "DATE", "in"), ("C", "TYPE", "in"),
    ("D", "LOT INTERNE", "in"), ("E", "EMPL. SOURCE", "in"),
    ("F", "EMPL. DESTINATION", "in"), ("G", "QUANTITE (KG)", "in"),
    ("H", "MOTIF / DESTINATION", "in"), ("I", "OBSERVATION", "in"),
    ("J", "LOT EXTERNE", "calc"), ("K", "ESPECE", "calc"),
    ("L", "PRODUCTEUR", "calc"), ("M", "CLASSE SOURCE", "calc"),
    ("N", "CLASSE DEST.", "calc"), ("O", "VALIDE", "calc"),
    ("P", "DISPO AVANT (KG)", "calc"), ("Q", "STATUT", "calc"),
    ("R", "IMPACT STOCK", "calc"), ("S", "CONTROLE / ACTION", "calc"),
]
OPS_W = {"A": 13, "B": 11, "C": 15, "D": 14, "E": 16, "F": 17, "G": 13,
         "H": 24, "I": 24, "J": 13, "K": 14, "L": 22, "M": 12, "N": 12,
         "O": 8, "P": 15, "Q": 21, "R": 10, "S": 46}


def build_operations(wb):
    ws = wb.create_sheet("OPERATIONS")
    r0, r1 = FIRST, FIRST + N_OPS - 1
    title_bar(ws, "OPERATIONS DE STOCK  -  RECEPTION / TRANSFERT / "
              "CONSOMMATION / AJUSTEMENT",
              "RECEPTION = entree sur un emplacement.  TRANSFERT = sortie d'un "
              "emplacement + entree sur un autre (meme quantite).  "
              "CONSOMMATION = sortie de la seule quantite consommee.  "
              "AJUSTEMENT = correction d'inventaire sur l'emplacement de "
              "destination : quantite signee (negative = perte) et MOTIF "
              "obligatoire.", "S")
    section(ws, 3, "Saisie minimale : DATE, TYPE, LOT INTERNE, "
            "EMPLACEMENT(S), QUANTITE. Le reste est deduit du lot.", 19)
    headers(ws, HDR, OPS_COLS, OPS_W)
    body_style(ws, OPS_COLS, r0, r1,
               fmts={"B": NUM_DATE, "G": NUM_KG0, "P": NUM_KG0, "O": NUM_INT,
                     "R": NUM_INT})
    ws.freeze_panes = "C%d" % FIRST

    fill_col(ws, "A", r0, r1,
             '=IF($C{r}="","","OP-"&TEXT(ROW()-%d,"000"))' % HDR)
    fill_col(ws, "J", r0, r1,
             '=IF($D{r}="","",IFERROR(INDEX(LOT_EXTE,MATCH($D{r},LOT_KEY,0)),'
             '""))')
    fill_col(ws, "K", r0, r1,
             '=IF($D{r}="","",IFERROR(INDEX(LOT_PRODE,MATCH($D{r},LOT_KEY,0)),'
             '"LOT NON REFERENCE"))')
    fill_col(ws, "L", r0, r1,
             '=IF($D{r}="","",IFERROR(INDEX(LOT_FOURNE,'
             'MATCH($D{r},LOT_KEY,0)),""))')
    fill_col(ws, "M", r0, r1,
             '=IF($E{r}="","",IFERROR(INDEX(EMPL_CLASSE,'
             'MATCH($E{r},EMPL_KEY,0)),"?"))')
    fill_col(ws, "N", r0, r1,
             '=IF($F{r}="","",IFERROR(INDEX(EMPL_CLASSE,'
             'MATCH($F{r},EMPL_KEY,0)),"?"))')
    fill_col(ws, "O", r0, r1,
             '=IF($C{r}="",0,IF(OR($B{r}="",$D{r}=""),0,'
             'IF($C{r}="AJUSTEMENT",'
             'IF(OR(NOT(ISNUMBER($G{r})),$G{r}=0,$F{r}="",$H{r}=""),0,1),'
             'IF(OR(NOT(ISNUMBER($G{r})),$G{r}<=0),0,'
             'IF($C{r}="RECEPTION",IF($F{r}="",0,1),'
             'IF($C{r}="CONSOMMATION",IF($E{r}="",0,1),'
             'IF(OR($E{r}="",$F{r}="",$E{r}=$F{r}),0,1)))))))')
    fill_col(ws, "P", r0, r1,
             '=IF(OR($O{r}=0,$C{r}="RECEPTION"),"",'
             'IF($C{r}="AJUSTEMENT",IF($G{r}>0,"",'
             'SUMIFS(MVT_BRUT,MVT_LOT,$D{r},MVT_EMPL,$F{r})-$G{r}),'
             'SUMIFS(MVT_BRUT,MVT_LOT,$D{r},MVT_EMPL,$E{r})+$G{r}))')
    fill_col(ws, "Q", r0, r1,
             '=IF($C{r}="","",IF($O{r}=0,"DONNEES MANQUANTES",'
             'IF($C{r}="RECEPTION","OK",'
             'IF(AND($C{r}="AJUSTEMENT",$G{r}>0),"OK",'
             'IF($P{r}<=0,"EMPLACEMENT INCOHERENT",'
             'IF(ABS($G{r})>$P{r}+0.001,"STOCK INSUFFISANT","OK"))))))')
    fill_col(ws, "R", r0, r1, '=IF($Q{r}="OK",1,0)')
    fill_col(ws, "S", r0, r1,
             '=IF($C{r}="","",IF($O{r}=0,"A COMPLETER : "'
             '&IF($B{r}="","DATE; ","")&IF($D{r}="","LOT INTERNE; ","")'
             '&IF(NOT(ISNUMBER($G{r})),"QUANTITE; ",'
             'IF(AND($C{r}="AJUSTEMENT",$G{r}=0),"QUANTITE <> 0; ",'
             'IF(AND($C{r}<>"AJUSTEMENT",$G{r}<=0),"QUANTITE > 0; ","")))'
             '&IF(AND($C{r}="AJUSTEMENT",$H{r}=""),"MOTIF; ","")'
             '&IF(AND(OR($C{r}="TRANSFERT",$C{r}="CONSOMMATION"),$E{r}=""),'
             '"EMPL. SOURCE; ","")'
             '&IF(AND($C{r}<>"CONSOMMATION",$F{r}=""),"EMPL. DESTINATION; ","")'
             '&IF(AND($C{r}="TRANSFERT",$E{r}<>"",$E{r}=$F{r}),'
             '"SOURCE = DESTINATION; ",""),'
             'IF($Q{r}="EMPLACEMENT INCOHERENT","Le lot "&$D{r}&" n\'existe pas'
             ' sur cet emplacement (stock 0). Verifier l\'emplacement.",'
             'IF($Q{r}="STOCK INSUFFISANT","Disponible avant operation : "'
             '&TEXT($P{r},"#,##0.00")&" kg. Ligne NON comptee en stock.",'
             'IF(COUNTIFS($B$%d:$B$%d,$B{r},$C$%d:$C$%d,$C{r},'
             '$D$%d:$D$%d,$D{r},$G$%d:$G$%d,$G{r})>1,'
             '"Doublon possible (meme date, type, lot et quantite).",'
             'IF($K{r}="LOT NON REFERENCE",'
             '"Creer d\'abord la fiche du lot dans l\'onglet LOTS.",""))))))'
             % (r0, r1, r0, r1, r0, r1, r0, r1))

    dv(ws, "C{0}:C{1}".format(r0, r1), ENUM_TYPEOP, strict=True,
       prompt="RECEPTION, TRANSFERT, CONSOMMATION ou AJUSTEMENT")
    dv(ws, "D{0}:D{1}".format(r0, r1), "L_LOTS", strict=True,
       prompt="Choisir un lot interne deja cree dans l'onglet LOTS.")
    dv(ws, "E{0}:E{1}".format(r0, r1), "L_EMPL", strict=True,
       prompt="Laisser vide pour une RECEPTION ou un AJUSTEMENT.")
    dv(ws, "F{0}:F{1}".format(r0, r1), "L_EMPL", strict=True,
       prompt="Laisser vide pour une CONSOMMATION.")
    dv(ws, "G{0}:G{1}".format(r0, r1), "0", kind="decimal", op="notEqual",
       strict=True,
       prompt="Quantite > 0. Seul un AJUSTEMENT accepte une valeur negative.")
    cf_status(ws, "Q%d:Q%d" % (r0, r1))
    return ws
# ============================================================================
# SOUS_TRAITANCE  (en-tete d'envoi)
# ============================================================================
SUITE = ('IF($U{r}=0,"EN COURS",IF($T{r}>$G{r}+0.001,"BLOQUE",'
         'IF(AND($V{r}>P_TOL,$L{r}=""),"A JUSTIFIER","OK")))')

ST_COLS = [
    ("A", "ID ST", "calc"), ("B", "DATE ENVOI", "in"),
    ("C", "SOUS-TRAITANT", "in"), ("D", "SOURCE", "in"),
    ("E", "LOT SOURCE (INTERNE)", "in"), ("F", "EMPL. SOURCE", "in"),
    ("G", "QTE ENVOYEE (KG)", "in"), ("H", "QUALITE ENTREE", "in"),
    ("I", "MOULE ENTREE", "in"), ("J", "HISTAMINE ENTREE (PPM)", "in"),
    ("K", "DATE RETOUR", "in"), ("L", "JUSTIFICATION ECART", "in"),
    ("M", "MOTIF ECART", "in"), ("N", "OBSERVATION", "in"),
    ("O", "LOT EXTERNE", "calc"), ("P", "ESPECE", "calc"),
    ("Q", "PRODUCTEUR", "calc"), ("R", "VALIDE", "calc"),
    ("S", "DISPO AVANT (KG)", "calc"), ("T", "TOTAL SORTIE (KG)", "calc"),
    ("U", "NB RESULTATS", "calc"), ("V", "ECART A JUSTIFIER (KG)", "calc"),
    ("W", "RENDEMENT", "calc"), ("X", "STATUT", "calc"),
    ("Y", "IMPACT STOCK", "calc"), ("Z", "CONTROLE / ACTION", "calc"),
    ("AA", "EN COURS (KG)", "calc"), ("AB", "REL. RECHERCHE", "calc"),
    ("AC", "RANG RECHERCHE", "calc"),
]
ST_W = {"A": 10, "B": 11, "C": 22, "D": 16, "E": 17, "F": 15, "G": 15,
        "H": 14, "I": 14, "J": 15, "K": 11, "L": 30, "M": 16, "N": 22,
        "O": 13, "P": 14, "Q": 22, "R": 8, "S": 15, "T": 15, "U": 11,
        "V": 16, "W": 11, "X": 21, "Y": 11, "Z": 46, "AA": 13, "AB": 12,
        "AC": 12}


def build_st(wb):
    ws = wb.create_sheet("SOUS_TRAITANCE")
    r0, r1 = FIRST, FIRST + N_ST - 1
    title_bar(ws, "SOUS-TRAITANCE  -  ENVOIS",
              "SOURCE = FOURNISSEUR : la marchandise va directement du "
              "fournisseur au sous-traitant, aucun stock OCEAMIC n'est "
              "consomme (ni emplacement ni reception a saisir).   "
              "SOURCE = STOCK EXISTANT : la quantite envoyee est retiree de "
              "l'emplacement indique.   Un envoi sans resultat reste EN COURS "
              "- ce n'est pas une erreur.", "AC")
    section(ws, 3, "Etape 1 : enregistrer l'envoi.  Etape 2 (plus tard) : "
            "saisir les resultats dans l'onglet ST_RESULTATS.", 29)
    headers(ws, HDR, ST_COLS, ST_W)
    body_style(ws, ST_COLS, r0, r1,
               fmts={"B": NUM_DATE, "K": NUM_DATE, "G": NUM_KG0, "S": NUM_KG0,
                     "T": NUM_KG0, "V": NUM_KG0, "AA": NUM_KG0, "W": NUM_PCT,
                     "U": NUM_INT, "R": NUM_INT, "Y": NUM_INT})
    ws.freeze_panes = "C%d" % FIRST

    fill_col(ws, "A", r0, r1,
             '=IF($D{r}="","","ST-"&TEXT(ROW()-%d,"000"))' % HDR)
    fill_col(ws, "O", r0, r1,
             '=IF($E{r}="","",IFERROR(INDEX(LOT_EXTE,MATCH($E{r},LOT_KEY,0)),'
             '""))')
    fill_col(ws, "P", r0, r1,
             '=IF($E{r}="","",IFERROR(INDEX(LOT_PRODE,MATCH($E{r},LOT_KEY,0)),'
             '"LOT NON REFERENCE"))')
    fill_col(ws, "Q", r0, r1,
             '=IF($E{r}="","",IFERROR(INDEX(LOT_FOURNE,'
             'MATCH($E{r},LOT_KEY,0)),""))')
    fill_col(ws, "R", r0, r1,
             '=IF($D{r}="",0,IF(OR($B{r}="",$C{r}="",$E{r}="",'
             'NOT(ISNUMBER($G{r})),$G{r}<=0),0,'
             'IF($D{r}="STOCK EXISTANT",IF($F{r}="",0,1),1)))')
    fill_col(ws, "S", r0, r1,
             '=IF(OR($R{r}=0,$D{r}<>"STOCK EXISTANT"),"",'
             'SUMIFS(MVT_BRUT,MVT_LOT,$E{r},MVT_EMPL,$F{r})+$G{r})')
    fill_col(ws, "T", r0, r1,
             '=IF($A{r}="","",SUMIFS(RES_QTEVAL,RES_IDST,$A{r}))')
    fill_col(ws, "U", r0, r1, '=IF($A{r}="","",COUNTIFS(RES_IDST,$A{r}))')
    fill_col(ws, "V", r0, r1, '=IF(OR($A{r}="",$U{r}=0),"",$G{r}-$T{r})')
    fill_col(ws, "W", r0, r1,
             '=IF(OR($A{r}="",$U{r}=0,NOT(ISNUMBER($G{r})),$G{r}<=0),"",'
             '$T{r}/$G{r})')
    fill_col(ws, "X", r0, r1,
             '=IF($D{r}="","",IF($R{r}=0,"DONNEES MANQUANTES",'
             'IF($D{r}="STOCK EXISTANT",IF($S{r}<=0,'
             '"EMPLACEMENT INCOHERENT",IF($G{r}>$S{r}+0.001,'
             '"STOCK INSUFFISANT",' + SUITE + ')),' + SUITE + ')))')
    fill_col(ws, "Y", r0, r1,
             '=IF(OR($X{r}="",$X{r}="DONNEES MANQUANTES",'
             '$X{r}="EMPLACEMENT INCOHERENT",$X{r}="STOCK INSUFFISANT"),0,1)')
    fill_col(ws, "Z", r0, r1,
             '=IF($D{r}="","",IF($R{r}=0,"A COMPLETER : "'
             '&IF($B{r}="","DATE ENVOI; ","")'
             '&IF($C{r}="","SOUS-TRAITANT; ","")'
             '&IF($E{r}="","LOT SOURCE; ","")'
             '&IF(OR(NOT(ISNUMBER($G{r})),$G{r}<=0),"QTE ENVOYEE > 0; ","")'
             '&IF(AND($D{r}="STOCK EXISTANT",$F{r}=""),"EMPL. SOURCE; ",""),'
             'IF($X{r}="EMPLACEMENT INCOHERENT","Le lot "&$E{r}&" n\'existe '
             'pas sur "&$F{r}&" (stock 0).",'
             'IF($X{r}="STOCK INSUFFISANT","Disponible avant envoi : "'
             '&TEXT($S{r},"#,##0.00")&" kg. Ligne NON comptee en stock.",'
             'IF($X{r}="BLOQUE","Total des sorties ("&TEXT($T{r},"#,##0.00")'
             '&" kg) superieur a la quantite envoyee.",'
             'IF($X{r}="EN COURS","Envoi enregistre, resultats non connus. '
             'A completer dans ST_RESULTATS.",'
             'IF($X{r}="A JUSTIFIER","Ecart de "&TEXT($V{r},"#,##0.00")'
             '&" kg : renseigner JUSTIFICATION ECART (colonne L).",'
             'IF($P{r}="LOT NON REFERENCE",'
             '"Creer d\'abord la fiche du lot dans l\'onglet LOTS.",""))))))))')
    fill_col(ws, "AA", r0, r1, '=IF(OR($Y{r}=0,$U{r}>0),0,$G{r})')
    fill_col(ws, "AB", r0, r1,
             '=IF($R{r}=0,0,IF(OR($E{r}=R_LOT,'
             'COUNTIFS(RES_IDST,$A{r},RES_LOTEFF,R_LOT)>0),1,0))')
    fill_col(ws, "AC", r0, r1,
             '=IF($AB{r}=0,"",COUNTIF($AB$%d:$AB{r},1))' % r0)

    dv(ws, "C{0}:C{1}".format(r0, r1), "L_STRAIT")
    dv(ws, "D{0}:D{1}".format(r0, r1), ENUM_SRCST, strict=True,
       prompt="FOURNISSEUR = livraison directe (aucun stock consomme). "
              "STOCK EXISTANT = prelevement sur stock.")
    dv(ws, "E{0}:E{1}".format(r0, r1), "L_LOTS", strict=True)
    dv(ws, "F{0}:F{1}".format(r0, r1), "L_EMPL", strict=True,
       prompt="A remplir uniquement si SOURCE = STOCK EXISTANT.")
    dv(ws, "G{0}:G{1}".format(r0, r1), "0", kind="decimal", op="greaterThan",
       strict=True)
    dv(ws, "H{0}:H{1}".format(r0, r1), "L_QUALITE")
    dv(ws, "I{0}:I{1}".format(r0, r1), "L_MOULE")
    dv(ws, "M{0}:M{1}".format(r0, r1), "L_MOTIF")
    cf_status(ws, "X%d:X%d" % (r0, r1))
    ws.conditional_formatting.add(
        "V%d:V%d" % (r0, r1),
        CellIsRule(operator="greaterThan", formula=["0"],
                   font=Font(name=FONT, size=9, bold=True, color=G_WARN[1])))
    return ws


# ============================================================================
# ST_RESULTATS
# ============================================================================
RES_COLS = [
    ("A", "ID RESULTAT", "calc"), ("B", "ID ST", "in"),
    ("C", "DATE RESULTAT", "in"), ("D", "RESULTAT / ETAT", "in"),
    ("E", "LOT RESULTAT (INTERNE)", "in"), ("F", "QUANTITE (KG)", "in"),
    ("G", "QUALITE SORTIE", "in"), ("H", "MOULE SORTIE", "in"),
    ("I", "HISTAMINE (PPM)", "in"), ("J", "EMPL. DESTINATION", "in"),
    ("K", "OBSERVATION", "in"), ("L", "LOT SOURCE", "calc"),
    ("M", "SOUS-TRAITANT", "calc"), ("N", "LOT RESULTAT EFFECTIF", "calc"),
    ("O", "LOT EXTERNE", "calc"), ("P", "CLASSE DEST.", "calc"),
    ("Q", "VALIDE", "calc"), ("R", "QTE VALIDE (KG)", "calc"),
    ("S", "STATUT ST", "calc"), ("T", "STATUT", "calc"),
    ("U", "IMPACT STOCK", "calc"), ("V", "CONTROLE / ACTION", "calc"),
    ("W", "DATE EFFECTIVE", "calc"), ("X", "LOT LIE", "calc"),
    ("Y", "RANG LIEN", "calc"), ("Z", "REL. RECHERCHE", "calc"),
    ("AA", "RANG RECHERCHE", "calc"),
]
RES_W = {"A": 12, "B": 10, "C": 12, "D": 14, "E": 18, "F": 13, "G": 14,
         "H": 14, "I": 13, "J": 17, "K": 22, "L": 12, "M": 20, "N": 18,
         "O": 13, "P": 12, "Q": 8, "R": 13, "S": 20, "T": 20, "U": 11,
         "V": 44, "W": 13, "X": 13, "Y": 10, "Z": 12, "AA": 12}


def build_res(wb):
    ws = wb.create_sheet("ST_RESULTATS")
    r0, r1 = FIRST, FIRST + N_RES - 1
    title_bar(ws, "SOUS-TRAITANCE  -  RESULTATS",
              "Un envoi peut produire plusieurs resultats (etats, qualites ou "
              "moules differents). Chaque resultat entre en stock a "
              "l'emplacement de destination.   LOT RESULTAT laisse vide = on "
              "conserve le lot source.", "AA")
    section(ws, 3, "Choisir l'ID ST, puis saisir uniquement : quantite, "
            "destination, et ce qui est connu (etat, qualite, moule).", 27)
    headers(ws, HDR, RES_COLS, RES_W)
    body_style(ws, RES_COLS, r0, r1,
               fmts={"C": NUM_DATE, "W": NUM_DATE, "F": NUM_KG0,
                     "R": NUM_KG0, "Q": NUM_INT, "U": NUM_INT})
    ws.freeze_panes = "C%d" % FIRST

    fill_col(ws, "A", r0, r1,
             '=IF($B{r}="","","RES-"&TEXT(ROW()-%d,"000"))' % HDR)
    fill_col(ws, "L", r0, r1,
             '=IF($B{r}="","",IFERROR(INDEX(ST_LOT,MATCH($B{r},ST_ID,0)),""))')
    fill_col(ws, "M", r0, r1,
             '=IF($B{r}="","",IFERROR(INDEX(ST_STRAIT,MATCH($B{r},ST_ID,0)),'
             '""))')
    fill_col(ws, "N", r0, r1, '=IF($B{r}="","",IF($E{r}="",$L{r},$E{r}))')
    fill_col(ws, "O", r0, r1,
             '=IF($N{r}="","",IFERROR(INDEX(LOT_EXTE,MATCH($N{r},LOT_KEY,0)),'
             '""))')
    fill_col(ws, "P", r0, r1,
             '=IF($J{r}="","",IFERROR(INDEX(EMPL_CLASSE,'
             'MATCH($J{r},EMPL_KEY,0)),"?"))')
    fill_col(ws, "Q", r0, r1,
             '=IF($B{r}="",0,IF(OR(ISNA(MATCH($B{r},ST_ID,0)),'
             'NOT(ISNUMBER($F{r})),$F{r}<=0,$J{r}="",$N{r}=""),0,1))')
    fill_col(ws, "R", r0, r1, '=IF($Q{r}=0,0,$F{r})')
    fill_col(ws, "S", r0, r1,
             '=IF($B{r}="","",IFERROR(INDEX(ST_STATUT,MATCH($B{r},ST_ID,0)),'
             '""))')
    fill_col(ws, "T", r0, r1,
             '=IF($B{r}="","",IF($Q{r}=0,"DONNEES MANQUANTES",'
             'IF(OR($S{r}="DONNEES MANQUANTES",$S{r}="STOCK INSUFFISANT",'
             '$S{r}="EMPLACEMENT INCOHERENT",$S{r}="BLOQUE"),"BLOQUE","OK")))')
    fill_col(ws, "U", r0, r1, '=IF($T{r}="OK",1,0)')
    fill_col(ws, "V", r0, r1,
             '=IF($B{r}="","",IF($Q{r}=0,"A COMPLETER : "'
             '&IF(ISNA(MATCH($B{r},ST_ID,0)),"ID ST valide; ","")'
             '&IF(OR(NOT(ISNUMBER($F{r})),$F{r}<=0),"QUANTITE > 0; ","")'
             '&IF($J{r}="","EMPL. DESTINATION; ",""),'
             'IF($T{r}="BLOQUE","Envoi parent en anomalie ("&$S{r}&") : '
             'corriger l\'onglet SOUS_TRAITANCE.",'
             'IF(AND($E{r}<>"",ISNA(MATCH($E{r},LOT_KEY,0))),'
             '"Creer la fiche du lot resultat dans l\'onglet LOTS.",""))))')
    fill_col(ws, "W", r0, r1,
             '=IF($B{r}="","",IF($C{r}<>"",$C{r},'
             'IFERROR(INDEX(ST_DATERET,MATCH($B{r},ST_ID,0)),"")))')
    fill_col(ws, "X", r0, r1,
             '=IF($Q{r}=0,"",IF($L{r}=R_LOT,IF($N{r}<>R_LOT,$N{r},""),'
             'IF($N{r}=R_LOT,$L{r},"")))')
    for r in range(r0, r1 + 1):
        ws["Y%d" % r] = ('=IF($X{r}="","",IF(COUNTIF($X${h}:$X{p},$X{r})>0,"",'
                         'COUNT($Y${h}:$Y{p})+1))').format(r=r, h=HDR, p=r - 1)
    fill_col(ws, "Z", r0, r1,
             '=IF($Q{r}=0,0,IF(OR($L{r}=R_LOT,$N{r}=R_LOT),1,0))')
    fill_col(ws, "AA", r0, r1,
             '=IF($Z{r}=0,"",COUNTIF($Z$%d:$Z{r},1))' % r0)

    dv(ws, "B{0}:B{1}".format(r0, r1), "L_IDST", strict=True,
       prompt="Choisir l'envoi concerne (onglet SOUS_TRAITANCE).")
    dv(ws, "D{0}:D{1}".format(r0, r1), "L_ETAT")
    dv(ws, "E{0}:E{1}".format(r0, r1), "L_LOTS", strict=True,
       prompt="Laisser vide pour conserver le lot source.")
    dv(ws, "F{0}:F{1}".format(r0, r1), "0", kind="decimal", op="greaterThan",
       strict=True)
    dv(ws, "G{0}:G{1}".format(r0, r1), "L_QUALITE")
    dv(ws, "H{0}:H{1}".format(r0, r1), "L_MOULE")
    dv(ws, "J{0}:J{1}".format(r0, r1), "L_EMPL", strict=True)
    cf_status(ws, "T%d:T%d" % (r0, r1))
    return ws


# ============================================================================
# QUALITE
# ============================================================================
QC_COLS = [
    ("A", "ID CONTROLE", "calc"), ("B", "DATE", "in"),
    ("C", "LOT INTERNE", "in"), ("D", "EMPLACEMENT", "in"),
    ("E", "TYPE DE CONTROLE", "in"), ("F", "TEMPERATURE (C)", "in"),
    ("G", "HISTAMINE (PPM)", "in"), ("H", "QUALITE", "in"),
    ("I", "MOULE", "in"), ("J", "DEFAUTS", "in"), ("K", "DECISION", "in"),
    ("L", "CONTROLEUR", "in"), ("M", "OBSERVATION", "in"),
    ("N", "LOT EXTERNE", "calc"), ("O", "ESPECE", "calc"),
    ("P", "CONSERVATION", "calc"), ("Q", "STOCK LOT / EMPL. (KG)", "calc"),
    ("R", "STATUT", "calc"), ("S", "ALERTE", "calc"),
    ("T", "DERNIER CONTROLE", "calc"), ("U", "CLE DERNIER", "calc"),
    ("V", "REL. RECHERCHE", "calc"), ("W", "RANG RECHERCHE", "calc"),
]
QC_W = {"A": 12, "B": 11, "C": 14, "D": 16, "E": 20, "F": 14, "G": 15,
        "H": 20, "I": 26, "J": 26, "K": 18, "L": 14, "M": 24, "N": 13,
        "O": 14, "P": 14, "Q": 17, "R": 21, "S": 46, "T": 12, "U": 13,
        "V": 12, "W": 12}


def build_qualite(wb):
    ws = wb.create_sheet("QUALITE")
    r0, r1 = FIRST, FIRST + N_QC - 1
    title_bar(ws, "CONTROLES QUALITE",
              "Un controle qualite ne modifie JAMAIS le stock. Saisir "
              "uniquement les mesures reellement effectuees : aucun champ "
              "n'est obligatoire au-dela du lot, de la date et d'au moins une "
              "information de controle.   QUALITE et MOULE acceptent la "
              "saisie libre (ex. A=20 B=40 C=40  /  20/24=14 26/30=55).", "W")
    section(ws, 3, "Le resultat appartient au lot physiquement controle. Un "
            "lot lie reste tracable mais n'herite pas du resultat.", 23)
    headers(ws, HDR, QC_COLS, QC_W)
    body_style(ws, QC_COLS, r0, r1,
               fmts={"B": NUM_DATE, "F": '0.0', "G": '#,##0',
                     "Q": NUM_KG0, "T": NUM_INT, "V": NUM_INT})
    ws.freeze_panes = "C%d" % FIRST

    fill_col(ws, "A", r0, r1,
             '=IF($C{r}="","","QC-"&TEXT(ROW()-%d,"000"))' % HDR)
    fill_col(ws, "N", r0, r1,
             '=IF($C{r}="","",IFERROR(INDEX(LOT_EXTE,MATCH($C{r},LOT_KEY,0)),'
             '""))')
    fill_col(ws, "O", r0, r1,
             '=IF($C{r}="","",IFERROR(INDEX(LOT_PRODE,MATCH($C{r},LOT_KEY,0)),'
             '"LOT NON REFERENCE"))')
    fill_col(ws, "P", r0, r1,
             '=IF($C{r}="","",IFERROR(INDEX(LOT_CONSERVE,'
             'MATCH($C{r},LOT_KEY,0)),""))')
    fill_col(ws, "Q", r0, r1,
             '=IF(OR($C{r}="",$D{r}=""),"",'
             'SUMIFS(MVT_NET,MVT_LOT,$C{r},MVT_EMPL,$D{r}))')
    fill_col(ws, "R", r0, r1,
             '=IF($C{r}="","",IF(OR($B{r}="",AND($F{r}="",$G{r}="",$H{r}="",'
             '$I{r}="",$J{r}="",$K{r}="")),"DONNEES MANQUANTES",'
             'IF(AND($D{r}<>"",$Q{r}<=0),"EMPLACEMENT INCOHERENT","OK")))')
    fill_col(ws, "S", r0, r1,
             '=IF($C{r}="","",IF(AND(ISNUMBER($G{r}),$G{r}>P_HIST),'
             '"HISTAMINE "&TEXT($G{r},"#,##0")&" PPM > seuil "'
             '&TEXT(P_HIST,"#,##0")&". ","")'
             '&IF(AND(ISNUMBER($F{r}),$P{r}="CONGELE",$F{r}>P_TCONG),'
             '"TEMPERATURE "&TEXT($F{r},"0.0")&" C hors seuil congele. ","")'
             '&IF(AND(ISNUMBER($F{r}),$P{r}="FRAIS",$F{r}>P_TFRAIS),'
             '"TEMPERATURE "&TEXT($F{r},"0.0")&" C hors seuil frais. ","")'
             '&IF($K{r}="BLOQUER","Lot bloque qualite : stock signale mais '
             'inchange. ","")'
             '&IF($R{r}="EMPLACEMENT INCOHERENT",'
             '"Lot absent de cet emplacement : controle conserve, stock '
             'inchange. ","")'
             '&IF($R{r}="DONNEES MANQUANTES",'
             '"Renseigner la DATE et au moins une mesure. ",""))')
    fill_col(ws, "T", r0, r1,
             '=IF(OR($C{r}="",$B{r}=""),0,'
             'IF(COUNTIFS(QC_LOT,$C{r},QC_DATE,">"&$B{r})>0,0,'
             'IF(COUNTIFS($C$%d:$C{r},$C{r},$B$%d:$B{r},$B{r})'
             '=COUNTIFS(QC_LOT,$C{r},QC_DATE,$B{r}),1,0)))' % (r0, r0))
    fill_col(ws, "U", r0, r1, '=IF($T{r}=1,$C{r},"")')
    fill_col(ws, "V", r0, r1, '=IF($C{r}="",0,IF($C{r}=R_LOT,1,0))')
    fill_col(ws, "W", r0, r1,
             '=IF($V{r}=0,"",COUNTIF($V$%d:$V{r},1))' % r0)

    dv(ws, "C{0}:C{1}".format(r0, r1), "L_LOTS", strict=True)
    dv(ws, "D{0}:D{1}".format(r0, r1), "L_EMPL", strict=True,
       prompt="Optionnel. Sert a verifier la coherence lot / emplacement.")
    dv(ws, "E{0}:E{1}".format(r0, r1), "L_TYPECTRL")
    dv(ws, "H{0}:H{1}".format(r0, r1), "L_QUALITE")
    dv(ws, "I{0}:I{1}".format(r0, r1), "L_MOULE")
    dv(ws, "K{0}:K{1}".format(r0, r1), "L_DECISION")
    cf_status(ws, "R%d:R%d" % (r0, r1))
    cf_status(ws, "K%d:K%d" % (r0, r1))
    ws.conditional_formatting.add(
        "S%d:S%d" % (r0, r1),
        FormulaRule(formula=['LEN($S%d)>0' % r0],
                    fill=PatternFill("solid", fgColor=G_WARN[0]),
                    font=Font(name=FONT, size=9, color=G_WARN[1])))
    return ws
# ============================================================================
# MOUVEMENTS  -  grand livre unique, 100% formules
# ============================================================================
MV_COLS = [
    ("A", "N", "calc"), ("B", "ORIGINE", "calc"), ("C", "REFERENCE", "calc"),
    ("D", "DATE", "calc"), ("E", "SENS", "calc"), ("F", "OPERATION", "calc"),
    ("G", "LOT INTERNE", "calc"), ("H", "LOT EXTERNE", "calc"),
    ("I", "EMPLACEMENT", "calc"), ("J", "CLASSE", "calc"),
    ("K", "ESPECE", "calc"), ("L", "ETAT MATIERE", "calc"),
    ("M", "QTE DECLAREE (KG)", "calc"), ("N", "QTE EN STOCK (KG)", "calc"),
    ("O", "STATUT LIGNE", "calc"), ("P", "RANG RECHERCHE", "calc"),
]
MV_W = {"A": 7, "B": 16, "C": 12, "D": 11, "E": 10, "F": 24, "G": 14,
        "H": 13, "I": 16, "J": 11, "K": 14, "L": 14, "M": 16, "N": 16,
        "O": 21, "P": 12}


def build_mouvements(wb):
    ws = wb.create_sheet("MOUVEMENTS")
    title_bar(ws, "MOUVEMENTS  -  GRAND LIVRE UNIQUE DU STOCK",
              "Feuille entierement calculee : NE RIEN SAISIR ICI. Chaque "
              "operation valide y produit automatiquement sa ou ses lignes.  "
              "ENTREES - SORTIES +/- AJUSTEMENTS = STOCK ACTUEL.", "P")
    section(ws, 3, "QTE DECLAREE = mouvement declare (sert au calcul du "
            "disponible).   QTE EN STOCK = mouvement reellement compte "
            "(0 si la ligne est en anomalie).", 16)
    headers(ws, HDR, MV_COLS, MV_W)
    body_style(ws, MV_COLS, MV_B1, MV_END,
               fmts={"D": NUM_DATE, "M": NUM_KG, "N": NUM_KG, "A": NUM_INT})
    ws.freeze_panes = "C%d" % FIRST

    common = {
        "A": '=IF($M{r}=0,"",ROW()-%d)' % HDR,
        "H": '=IF($G{r}="","",IFERROR(INDEX(LOT_EXTE,'
             'MATCH($G{r},LOT_KEY,0)),""))',
        "J": '=IF($I{r}="","",IFERROR(INDEX(EMPL_CLASSE,'
             'MATCH($I{r},EMPL_KEY,0)),""))',
        "L": '=IF($G{r}="","",IF(IFERROR(INDEX(LOT_ETAT,'
             'MATCH($G{r},LOT_KEY,0)),"")="","",'
             'INDEX(LOT_ETAT,MATCH($G{r},LOT_KEY,0))))',
    }

    def block(dest0, src0, n, spec):
        for i in range(n):
            r = dest0 + i
            s = src0 + i
            for col, tpl in spec.items():
                ws["%s%d" % (col, r)] = tpl.format(r=r, s=s)
            for col, tpl in common.items():
                ws["%s%d" % (col, r)] = tpl.format(r=r, s=s)

    # --- bloc 1 : sorties issues de OPERATIONS ---------------------------
    block(MV_B1, FIRST, N_OPS, {
        "M": '=IF(OR(OPERATIONS!$O{s}=0,OPERATIONS!$C{s}="RECEPTION",'
             'OPERATIONS!$C{s}="AJUSTEMENT"),0,-OPERATIONS!$G{s})',
        "N": '=IF($M{r}=0,0,OPERATIONS!$R{s}*$M{r})',
        "B": '=IF($M{r}=0,"","OPERATIONS")',
        "C": '=IF($M{r}=0,"",OPERATIONS!$A{s})',
        "D": '=IF($M{r}=0,"",OPERATIONS!$B{s})',
        "E": '=IF($M{r}=0,"","SORTIE")',
        "F": '=IF($M{r}=0,"",OPERATIONS!$C{s})',
        "G": '=IF($M{r}=0,"",OPERATIONS!$D{s})',
        "I": '=IF($M{r}=0,"",OPERATIONS!$E{s})',
        "K": '=IF($M{r}=0,"",OPERATIONS!$K{s})',
        "O": '=IF($M{r}=0,"",OPERATIONS!$Q{s})',
    })
    # --- bloc 2 : entrees et ajustements issus de OPERATIONS -------------
    block(MV_B2, FIRST, N_OPS, {
        "M": '=IF(OR(OPERATIONS!$O{s}=0,OPERATIONS!$C{s}="CONSOMMATION"),0,'
             'OPERATIONS!$G{s})',
        "N": '=IF($M{r}=0,0,OPERATIONS!$R{s}*$M{r})',
        "B": '=IF($M{r}=0,"","OPERATIONS")',
        "C": '=IF($M{r}=0,"",OPERATIONS!$A{s})',
        "D": '=IF($M{r}=0,"",OPERATIONS!$B{s})',
        "E": '=IF($M{r}=0,"",IF($M{r}<0,"SORTIE","ENTREE"))',
        "F": '=IF($M{r}=0,"",OPERATIONS!$C{s})',
        "G": '=IF($M{r}=0,"",OPERATIONS!$D{s})',
        "I": '=IF($M{r}=0,"",OPERATIONS!$F{s})',
        "K": '=IF($M{r}=0,"",OPERATIONS!$K{s})',
        "O": '=IF($M{r}=0,"",OPERATIONS!$Q{s})',
    })
    # --- bloc 3 : sorties vers sous-traitance ----------------------------
    block(MV_B3, FIRST, N_ST, {
        "M": '=IF(OR(SOUS_TRAITANCE!$R{s}=0,'
             'SOUS_TRAITANCE!$D{s}<>"STOCK EXISTANT"),0,-SOUS_TRAITANCE!$G{s})',
        "N": '=IF($M{r}=0,0,SOUS_TRAITANCE!$Y{s}*$M{r})',
        "B": '=IF($M{r}=0,"","SOUS-TRAITANCE")',
        "C": '=IF($M{r}=0,"",SOUS_TRAITANCE!$A{s})',
        "D": '=IF($M{r}=0,"",SOUS_TRAITANCE!$B{s})',
        "E": '=IF($M{r}=0,"","SORTIE")',
        "F": '=IF($M{r}=0,"","ENVOI SOUS-TRAITANCE")',
        "G": '=IF($M{r}=0,"",SOUS_TRAITANCE!$E{s})',
        "I": '=IF($M{r}=0,"",SOUS_TRAITANCE!$F{s})',
        "K": '=IF($M{r}=0,"",SOUS_TRAITANCE!$P{s})',
        "O": '=IF($M{r}=0,"",SOUS_TRAITANCE!$X{s})',
    })
    # --- bloc 4 : entrees issues des resultats de sous-traitance ---------
    block(MV_B4, FIRST, N_RES, {
        "M": '=IF(ST_RESULTATS!$Q{s}=0,0,ST_RESULTATS!$F{s})',
        "N": '=IF($M{r}=0,0,ST_RESULTATS!$U{s}*$M{r})',
        "B": '=IF($M{r}=0,"","RESULTAT ST")',
        "C": '=IF($M{r}=0,"",ST_RESULTATS!$A{s})',
        "D": '=IF($M{r}=0,"",ST_RESULTATS!$W{s})',
        "E": '=IF($M{r}=0,"","ENTREE")',
        "F": '=IF($M{r}=0,"","RETOUR SOUS-TRAITANCE")',
        "G": '=IF($M{r}=0,"",ST_RESULTATS!$N{s})',
        "I": '=IF($M{r}=0,"",ST_RESULTATS!$J{s})',
        "K": '=IF($M{r}=0,"",IFERROR(INDEX(LOT_PRODE,'
             'MATCH($G{r},LOT_KEY,0)),""))',
        "O": '=IF($M{r}=0,"",ST_RESULTATS!$T{s})',
    })

    for r in range(MV_B1, MV_END + 1):
        ws["P%d" % r] = ('=IF($G{r}="","",IF($G{r}=R_LOT,'
                         'COUNTIF($G${f}:$G{r},R_LOT),""))').format(
                             r=r, f=MV_B1)

    cf_status(ws, "O%d:O%d" % (MV_B1, MV_END))
    ws.conditional_formatting.add(
        "N%d:N%d" % (MV_B1, MV_END),
        FormulaRule(formula=['AND($M%d<>0,$N%d=0)' % (MV_B1, MV_B1)],
                    fill=PatternFill("solid", fgColor=G_BAD[0]),
                    font=Font(name=FONT, size=9, color=G_BAD[1])))
    return ws


# ============================================================================
# STOCK  -  synthese
# ============================================================================
HDR_M = 14                      # ligne d'en-tete de la matrice
R0_M = HDR_M + 1                # premiere ligne de lot
R1_M = HDR_M + N_LOTS
RT_M = R1_M + 1                 # ligne TOTAL
SYN = RT_M + 3                  # premiere ligne du bloc de syntheses


def build_stock(wb):
    ws = wb.create_sheet("STOCK")
    title_bar(ws, "SYNTHESE DES STOCKS",
              "Reconstruit integralement a partir du grand livre MOUVEMENTS. "
              "Une case vide signifie zero.", "U")
    kpi = [
        ("STOCK TOTAL INTERNE (kg)", '=SUMIFS(MVT_NET,MVT_CLASSE,"INTERNE")'),
        ("STOCK TOTAL EXTERNE (kg)", '=SUMIFS(MVT_NET,MVT_CLASSE,"EXTERNE")'),
        ("STOCK TOTAL (kg)", '=SUM(MVT_NET)'),
        ("NB LOTS EN STOCK", '=COUNTIF($N$%d:$N$%d,">0")' % (R0_M, R1_M)),
        ("EN COURS CHEZ LES SOUS-TRAITANTS (kg)", '=SUM(ST_ENCOURS)'),
        ("STOCK BLOQUE QUALITE (kg)", '=SUM($T$%d:$T$%d)' % (R0_M, R1_M)),
        ("CONTROLE : ECART MATRICE / GRAND LIVRE (kg)",
         '=ROUND(SUM(MVT_NET)-SUM($N$%d:$N$%d),3)' % (R0_M, R1_M)),
    ]
    for i, (lab, f) in enumerate(kpi):
        r = 4 + i
        ws.merge_cells("A%d:C%d" % (r, r))
        cell(ws, "A%d" % r, lab, bold=(i >= 5), size=9, color=INK, fill=KPI_BG,
             border=BOX, align="left", indent=1)
        cell(ws, "D%d" % r, f, bold=True, size=10, color=NAVY, fill=KPI_BG,
             border=BOX, align="center", fmt=NUM_KG0 if i != 3 else NUM_INT)
    ws.conditional_formatting.add(
        "D9", CellIsRule(operator="greaterThan", formula=["0"],
                         fill=PatternFill("solid", fgColor=G_BAD[0]),
                         font=Font(name=FONT, size=10, bold=True,
                                   color=G_BAD[1])))
    ws.conditional_formatting.add(
        "D10", CellIsRule(operator="notEqual", formula=["0"],
                          fill=PatternFill("solid", fgColor=G_BAD[0]),
                          font=Font(name=FONT, size=10, bold=True,
                                    color=G_BAD[1])))
    ws.merge_cells("E10:K10")
    cell(ws, "E10", "Doit rester a 0. Sinon un lot present dans les "
         "mouvements n'a pas de fiche dans LOTS.", size=8, color=MUTED,
         align="left")

    section(ws, 12, "STOCK PAR LOT ET PAR EMPLACEMENT", 21)
    cell(ws, "A13", "CLASSE", bold=True, size=8, color=MUTED, align="center",
         border=BOX, fill=CALC_BG)
    for j in range(N_COLS_EMPL):
        L = get_column_letter(2 + j)
        ws["%s13" % L] = ('=IF(INDEX(EMPL_KEY,%d)="","",'
                          'IFERROR(INDEX(EMPL_CLASSE,%d),""))' % (j + 1, j + 1))
        cell(ws, "%s13" % L, None, bold=True, size=8, color=MUTED,
             align="center", border=BOX, fill=CALC_BG)
        ws["%s%d" % (L, HDR_M)] = ('=IF(INDEX(EMPL_KEY,%d)="","",'
                                   'INDEX(EMPL_KEY,%d))' % (j + 1, j + 1))
        cell(ws, "%s%d" % (L, HDR_M), None, bold=True, size=9, color=WHITE,
             fill=TEAL, align="center", wrap=True, border=BOX)
        ws.column_dimensions[L].width = 13
    cell(ws, "A%d" % HDR_M, "LOT INTERNE", bold=True, size=9, color=WHITE,
         fill=TEAL, align="center", border=BOX)
    ws.column_dimensions["A"].width = 15
    extra = [("N", "TOTAL LOT (KG)"), ("O", "DONT INTERNE"),
             ("P", "DONT EXTERNE"), ("Q", "EN COURS ST"), ("R", "LOT EXTERNE"),
             ("S", "DERNIERE DECISION"), ("T", "STOCK BLOQUE (KG)"),
             ("U", "ALERTE")]
    for L, lab in extra:
        cell(ws, "%s%d" % (L, HDR_M), lab, bold=True, size=9, color=WHITE,
             fill=MUTED, align="center", wrap=True, border=BOX)
        ws.column_dimensions[L].width = 15
    ws.column_dimensions["S"].width = 20
    ws.column_dimensions["U"].width = 22
    ws.row_dimensions[HDR_M].height = 28

    for r in range(R0_M, R1_M + 1):
        src = FIRST + (r - R0_M)
        ws["A%d" % r] = '=IF(LOTS!$A%d="","",LOTS!$A%d)' % (src, src)
        cell(ws, "A%d" % r, None, size=9, bold=True, color=INK, fill=WHITE,
             border=BOX)
        for j in range(N_COLS_EMPL):
            L = get_column_letter(2 + j)
            ws["%s%d" % (L, r)] = ('=IF(OR($A{r}="",{L}${h}=""),0,'
                                   'SUMIFS(MVT_NET,MVT_LOT,$A{r},MVT_EMPL,'
                                   '{L}${h}))').format(r=r, L=L, h=HDR_M)
            cell(ws, "%s%d" % (L, r), None, size=9, fill=WHITE, border=BOX,
                 align="right", fmt=NUM_KG)
        ws["N%d" % r] = '=SUM($B%d:$M%d)' % (r, r)
        ws["O%d" % r] = '=SUMPRODUCT(($B$13:$M$13="INTERNE")*$B%d:$M%d)' % (r, r)
        ws["P%d" % r] = '=SUMPRODUCT(($B$13:$M$13="EXTERNE")*$B%d:$M%d)' % (r, r)
        ws["Q%d" % r] = '=IF($A%d="",0,SUMIFS(ST_ENCOURS,ST_LOT,$A%d))' % (r, r)
        ws["R%d" % r] = ('=IF($A{r}="","",IFERROR(INDEX(LOT_EXTE,'
                         'MATCH($A{r},LOT_KEY,0)),""))').format(r=r)
        ws["S%d" % r] = ('=IF($A{r}="","",IFERROR(INDEX(LOT_DECISION,'
                         'MATCH($A{r},LOT_KEY,0)),""))').format(r=r)
        ws["T%d" % r] = '=IF($S%d="BLOQUER",MAX(0,$N%d),0)' % (r, r)
        ws["U%d" % r] = ('=IF($A{r}="","",IF($N{r}<-0.001,'
                         '"STOCK NEGATIF - ANOMALIE",'
                         'IF($T{r}>0,"BLOQUE QUALITE","")))').format(r=r)
        for L in ("N", "O", "P", "Q", "T"):
            cell(ws, "%s%d" % (L, r), None, size=9, bold=(L == "N"),
                 fill=CALC_BG, border=BOX, align="right", fmt=NUM_KG)
        for L in ("R", "S"):
            cell(ws, "%s%d" % (L, r), None, size=9, fill=CALC_BG, border=BOX,
                 align="center", color=MUTED)
        cell(ws, "U%d" % r, None, size=9, fill=CALC_BG, border=BOX,
             color=G_BAD[1], bold=True, align="center")

    cell(ws, "A%d" % RT_M, "TOTAL", bold=True, size=9, color=WHITE, fill=NAVY,
         align="center", border=BOX)
    for j in range(N_COLS_EMPL):
        L = get_column_letter(2 + j)
        ws["%s%d" % (L, RT_M)] = '=SUM(%s%d:%s%d)' % (L, R0_M, L, R1_M)
        cell(ws, "%s%d" % (L, RT_M), None, bold=True, size=9, color=WHITE,
             fill=NAVY, border=BOX, align="right", fmt=NUM_KG)
    for L in ("N", "O", "P", "Q", "T"):
        ws["%s%d" % (L, RT_M)] = '=SUM(%s%d:%s%d)' % (L, R0_M, L, R1_M)
        cell(ws, "%s%d" % (L, RT_M), None, bold=True, size=9, color=WHITE,
             fill=NAVY, border=BOX, align="right", fmt=NUM_KG)
    for L in ("R", "S", "U"):
        cell(ws, "%s%d" % (L, RT_M), "", fill=NAVY, border=BOX)

    ws.conditional_formatting.add(
        "B%d:N%d" % (R0_M, R1_M),
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=PatternFill("solid", fgColor=G_BAD[0]),
                   font=Font(name=FONT, size=9, bold=True, color=G_BAD[1])))
    for val, (bg, fg) in (("BLOQUER", G_BAD), ("REFUSER", G_BAD),
                          ("A RECONTROLER", G_WARN), ("ACCEPTER", G_OK)):
        ws.conditional_formatting.add("S%d:S%d" % (R0_M, R1_M), CellIsRule(
            operator="equal", formula=['"%s"' % val],
            fill=PatternFill("solid", fgColor=bg),
            font=Font(name=FONT, size=9, bold=True, color=fg)))

    # ------------------------------------------------------------------
    # Syntheses complementaires
    # ------------------------------------------------------------------
    section(ws, SYN - 1, "SYNTHESES  -  PAR ESPECE, PAR ETAT MATIERE, "
            "RECONCILIATION DU GRAND LIVRE", 21)
    trio = [("A", "B", "ESPECE", "L_PRODUIT_R", 12),
            ("D", "E", "ETAT MATIERE", "L_ETAT_R", 10)]
    for c1, c2, lab, rng, n in trio:
        cell(ws, "%s%d" % (c1, SYN), lab, bold=True, size=9, color=WHITE,
             fill=TEAL, align="center", border=BOX)
        cell(ws, "%s%d" % (c2, SYN), "STOCK (KG)", bold=True, size=9,
             color=WHITE, fill=TEAL, align="center", border=BOX)
        for i in range(n):
            r = SYN + 1 + i
            ws["%s%d" % (c1, r)] = ('=IF(INDEX(%s,%d)="","",INDEX(%s,%d))'
                                    % (rng, i + 1, rng, i + 1))
            crit = "MVT_PROD" if lab == "ESPECE" else "MVT_ETAT"
            ws["%s%d" % (c2, r)] = ('=IF(%s%d="","",SUMIFS(MVT_NET,%s,%s%d))'
                                    % (c1, r, crit, c1, r))
            cell(ws, "%s%d" % (c1, r), None, size=9, fill=WHITE, border=BOX)
            cell(ws, "%s%d" % (c2, r), None, size=9, fill=WHITE, border=BOX,
                 align="right", fmt=NUM_KG0)
        rr = SYN + 1 + n
        cell(ws, "%s%d" % (c1, rr), "TOTAL", bold=True, size=9, color=WHITE,
             fill=NAVY, border=BOX, align="center")
        ws["%s%d" % (c2, rr)] = '=SUM(%s%d:%s%d)' % (c2, SYN + 1, c2, rr - 1)
        cell(ws, "%s%d" % (c2, rr), None, bold=True, size=9, color=WHITE,
             fill=NAVY, border=BOX, align="right", fmt=NUM_KG0)

    cell(ws, "G%d" % SYN, "RECONCILIATION DU GRAND LIVRE", bold=True, size=9,
         color=WHITE, fill=TEAL, align="center", border=BOX)
    cell(ws, "H%d" % SYN, "KG", bold=True, size=9, color=WHITE, fill=TEAL,
         align="center", border=BOX)
    rec = [
        ("ENTREES (hors ajustements)",
         '=SUMIFS(MVT_NET,MVT_NET,">0",MVT_OP,"<>AJUSTEMENT")'),
        ("SORTIES (hors ajustements)",
         '=-SUMIFS(MVT_NET,MVT_NET,"<0",MVT_OP,"<>AJUSTEMENT")'),
        ("AJUSTEMENTS (signes)", '=SUMIFS(MVT_NET,MVT_OP,"AJUSTEMENT")'),
        ("STOCK THEORIQUE", '=$H%d-$H%d+$H%d' % (SYN + 1, SYN + 2, SYN + 3)),
        ("STOCK ACTUEL", '=SUM(MVT_NET)'),
        ("ECART", '=ROUND($H%d-$H%d,3)' % (SYN + 5, SYN + 4)),
        ("BILAN", '=IF(ROUND($H%d,3)=0,"BILAN OK","BILAN A CORRIGER")'
         % (SYN + 6)),
        ("Ecarts sous-traitance non justifies (info)",
         '=SUMIFS(ST_ECART,ST_STATUT,"A JUSTIFIER")'),
    ]
    for i, (lab, f) in enumerate(rec):
        r = SYN + 1 + i
        bold = lab in ("STOCK THEORIQUE", "STOCK ACTUEL", "BILAN")
        cell(ws, "G%d" % r, lab, bold=bold, size=9,
             color=INK if bold else MUTED, fill=WHITE, border=BOX,
             align="left", indent=1)
        cell(ws, "H%d" % r, f, bold=bold, size=9, color=INK, fill=WHITE,
             border=BOX, align="right" if lab != "BILAN" else "center",
             fmt=None if lab == "BILAN" else NUM_KG0)
    ws.column_dimensions["G"].width = 34
    ws.column_dimensions["H"].width = 16
    ws.conditional_formatting.add(
        "H%d" % (SYN + 7), CellIsRule(
            operator="equal", formula=['"BILAN OK"'],
            fill=PatternFill("solid", fgColor=G_OK[0]),
            font=Font(name=FONT, size=9, bold=True, color=G_OK[1])))
    ws.conditional_formatting.add(
        "H%d" % (SYN + 7), CellIsRule(
            operator="equal", formula=['"BILAN A CORRIGER"'],
            fill=PatternFill("solid", fgColor=G_BAD[0]),
            font=Font(name=FONT, size=9, bold=True, color=G_BAD[1])))
    ws.freeze_panes = "B%d" % R0_M
    return ws
# ============================================================================
# RECHERCHE
# ============================================================================
IDENT = [
    (8,  "LOT EXTERNE",              "LOT_EXT",     "LOT_EXTE"),
    (9,  "ESPECE",                   "LOT_PROD",    "LOT_PRODE"),
    (10, "PRODUCTEUR / FOURNISSEUR", "LOT_FOURN",   "LOT_FOURNE"),
    (11, "ORIGINE",                  "LOT_ORIG",    "LOT_ORIGE"),
    (12, "MATRICULE CAMION",         "LOT_IMMAT",   "LOT_IMMATE"),
    (13, "ETAT MATIERE",             "LOT_ETAT",    None),
    (14, "CONSERVATION",             "LOT_CONSERV", "LOT_CONSERVE"),
    (15, "QUALITE INITIALE",         "LOT_QUAL",    None),
    (16, "MOULE INITIAL",            "LOT_MOULE",   None),
    (17, "DATE CREATION",            "LOT_DATE",    None),
    (18, "LOT PARENT / SOURCE",      "LOT_PARENTE", None),
    (19, "DERNIERE DECISION QUALITE", "LOT_DECISION", None),
    (20, "OBSERVATION",              "LOT_OBS",     None),
    (21, "STATUT DE LA FICHE",       "LOT_STATUT",  None),
]
PARENT_CELL = "$C$18"

RECH_BLOCKS = [
    ("HISTORIQUE DES MOUVEMENTS DE STOCK", 25,
     ["N", "DATE", "OPERATION", "SENS", "EMPLACEMENT", "CLASSE",
      "QTE DECLAREE", "QTE EN STOCK", "STATUT", "REFERENCE"], "MVT_RANG"),
    ("SOUS-TRAITANCE  (envois lies a ce lot)", 12,
     ["ID ST", "DATE ENVOI", "SOUS-TRAITANT", "SOURCE", "EMPL. SOURCE",
      "QTE ENVOYEE", "TOTAL SORTIE", "ECART", "NB RESULTATS", "STATUT"],
     "ST_RANG"),
    ("RESULTATS DE SOUS-TRAITANCE", 12,
     ["ID ST", "DATE", "RESULTAT / ETAT", "LOT RESULTAT", "QUANTITE",
      "QUALITE", "MOULE", "DESTINATION", "CLASSE", "STATUT"], "RES_RANG"),
    ("HISTORIQUE QUALITE", 12,
     ["DATE", "EMPLACEMENT", "TEMPERATURE", "HISTAMINE (PPM)", "QUALITE",
      "MOULE", "DEFAUTS", "DECISION", "STATUT", "ALERTE"], "QC_RANG"),
    ("LOTS LIES PAR SOUS-TRAITANCE  (sans doublon)", 8,
     ["LOT LIE", "ESPECE", "ETAT MATIERE", "STOCK TOTAL (KG)", "RELATION",
      "", "", "", "", ""], "RES_RANGLIEN"),
    ("AUTRES LOTS INTERNES PORTANT LE MEME LOT EXTERNE", 8,
     ["LOT INTERNE", "LOT EXTERNE", "ESPECE", "ETAT MATIERE",
      "STOCK TOTAL (KG)", "", "", "", "", ""], "LOT_RANGEXT"),
]

SRC = {
    "MVT_RANG": ["MVT_NO", "MVT_DATE", "MVT_OP", "MVT_SENS", "MVT_EMPL",
                 "MVT_CLASSE", "MVT_BRUT", "MVT_NET", "MVT_STATUT",
                 "MVT_REF"],
    "ST_RANG": ["ST_ID", "ST_DATE", "ST_STRAIT", "ST_SOURCE", "ST_EMPL",
                "ST_QTE", "ST_TOTSORTIE", "ST_ECART", "ST_NBRES",
                "ST_STATUT"],
    "RES_RANG": ["RES_IDST", "RES_DATE", "RES_ETAT", "RES_LOTEFF",
                 "RES_QTE", "RES_QUAL", "RES_MOULE", "RES_DEST", None,
                 "RES_STATUT"],
    "QC_RANG": ["QC_DATE", "QC_EMPL", "QC_TEMP", "QC_HIST", "QC_QUAL",
                "QC_MOULE", "QC_DEF", "QC_DEC", "QC_STATUT", "QC_ALERTE"],
    "RES_RANGLIEN": ["RES_LOTLIE", None, None, None, None,
                     None, None, None, None, None],
    "LOT_RANGEXT": ["LOT_KEY", "LOT_EXTE", "LOT_PRODE", "LOT_ETAT",
                    "LOT_STOCK", None, None, None, None, None],
}
DATES = {"MVT_DATE", "ST_DATE", "RES_DATE", "QC_DATE"}
KGS = {"MVT_BRUT", "MVT_NET", "ST_QTE", "ST_TOTSORTIE", "ST_ECART",
       "RES_QTE", "LOT_STOCK"}


def build_recherche(wb):
    ws = wb.create_sheet("RECHERCHE")
    title_bar(ws, "RECHERCHE PAR LOT",
              "Saisissez un LOT INTERNE ou un LOT EXTERNE : identite, stock, "
              "historique complet, sous-traitance, qualite et lots lies "
              "s'affichent automatiquement.", "K")
    widths = {"A": 2, "B": 27, "C": 17, "D": 18, "E": 18, "F": 15, "G": 16,
              "H": 16, "I": 16, "J": 21, "K": 30}
    for k, v in widths.items():
        ws.column_dimensions[k].width = v

    cell(ws, "B4", "LOT A RECHERCHER", bold=True, size=11, color=NAVY)
    ws.merge_cells("C4:E4")
    cell(ws, "C4", None, bold=True, size=12, color="1B4F8F", fill=INPUT_BG,
         align="center", border=BOX)
    ws.row_dimensions[4].height = 24
    dv(ws, "C4", "L_LOTS", strict=False)
    cell(ws, "B5", "LOT INTERNE ANALYSE", bold=True, size=9, color=MUTED)
    ws.merge_cells("C5:E5")
    ws["C5"] = ('=IF($C$4="","",IF(ISNUMBER(MATCH($C$4,LOT_KEY,0)),$C$4,'
                'IFERROR(INDEX(LOT_KEY,MATCH($C$4,LOT_EXTE,0)),"")))')
    cell(ws, "C5", None, bold=True, size=11, color=NAVY, fill=CALC_BG,
         align="center", border=BOX)
    ws.merge_cells("G4:K5")
    cell(ws, "G4", "Un LOT EXTERNE peut correspondre a plusieurs lots "
         "internes : le premier est analyse ici, les autres sont listes en "
         "bas de page.", size=8, color=MUTED, align="left", wrap=True,
         indent=1)

    ws.merge_cells("A7:E7")
    cell(ws, "A7", "IDENTITE DU LOT  (reprise automatique - aucune re-saisie)",
         bold=True, size=10, color=NAVY, fill=TEAL_SOFT, align="left",
         indent=1)
    ws.merge_cells("G7:I7")
    cell(ws, "G7", "STOCK ACTUEL", bold=True, size=10, color=NAVY,
         fill=TEAL_SOFT, align="left", indent=1)

    for row, label, name, eff in IDENT:
        cell(ws, "B%d" % row, label, bold=True, size=9, color=MUTED,
             fill=CALC_BG, border=BOX, align="left", indent=1)
        ws.merge_cells("C%d:E%d" % (row, row))
        cell(ws, "C%d" % row, None, size=10, color=INK, fill=WHITE,
             border=BOX, align="left", indent=1,
             fmt=NUM_DATE if name == "LOT_DATE" else None)
        own = ('IF($C$5="","",IF(IFERROR(INDEX({n},MATCH($C$5,LOT_KEY,0)),"")'
               '="","",INDEX({n},MATCH($C$5,LOT_KEY,0))))'.format(n=name))
        if eff is None:
            ws["C%d" % row] = "=" + own
        else:
            ws["C%d" % row] = (
                '=IF($C$5="","",IF(IFERROR(INDEX({n},MATCH($C$5,LOT_KEY,0)),"")'
                '<>"",INDEX({n},MATCH($C$5,LOT_KEY,0)),'
                'IF(IFERROR(INDEX({e},MATCH($C$5,LOT_KEY,0)),"")="","",'
                'INDEX({e},MATCH($C$5,LOT_KEY,0))'
                '&"   (herite du lot "&{p}&")")))').format(n=name, e=eff,
                                                           p=PARENT_CELL)
    cf_status(ws, "C21")
    for val, (bg, fg) in (("BLOQUER", G_BAD), ("REFUSER", G_BAD),
                          ("A RECONTROLER", G_WARN), ("ACCEPTER", G_OK)):
        ws.conditional_formatting.add("C19", CellIsRule(
            operator="equal", formula=['"%s"' % val],
            fill=PatternFill("solid", fgColor=bg),
            font=Font(name=FONT, size=10, bold=True, color=fg)))

    for j, lab in enumerate(("EMPLACEMENT", "CLASSE", "STOCK (KG)")):
        L = "GHI"[j]
        cell(ws, "%s8" % L, lab, bold=True, size=9, color=WHITE, fill=TEAL,
             align="center", border=BOX)
    for i in range(N_COLS_EMPL):
        r = 9 + i
        ws["G%d" % r] = ('=IF(INDEX(EMPL_KEY,%d)="","",INDEX(EMPL_KEY,%d))'
                         % (i + 1, i + 1))
        ws["H%d" % r] = ('=IF($G%d="","",IFERROR(INDEX(EMPL_CLASSE,'
                         'MATCH($G%d,EMPL_KEY,0)),""))' % (r, r))
        ws["I%d" % r] = ('=IF(OR($G%d="",$C$5=""),"",'
                         'SUMIFS(MVT_NET,MVT_LOT,$C$5,MVT_EMPL,$G%d))'
                         % (r, r))
        cell(ws, "G%d" % r, None, size=9, fill=WHITE, border=BOX)
        cell(ws, "H%d" % r, None, size=9, fill=WHITE, border=BOX,
             align="center", color=MUTED)
        cell(ws, "I%d" % r, None, size=9, fill=WHITE, border=BOX,
             align="right", fmt=NUM_KG)
    ws.merge_cells("G21:H21")
    cell(ws, "G21", "STOCK TOTAL DU LOT", bold=True, size=9, color=WHITE,
         fill=NAVY, border=BOX, align="left", indent=1)
    ws["I21"] = '=IF($C$5="","",SUM($I$9:$I$20))'
    cell(ws, "I21", None, bold=True, size=10, color=WHITE, fill=NAVY,
         border=BOX, align="right", fmt=NUM_KG0)
    ws.merge_cells("G22:H22")
    cell(ws, "G22", "EN COURS CHEZ SOUS-TRAITANT", bold=True, size=9,
         color=NAVY, fill=KPI_BG, border=BOX, align="left", indent=1)
    ws["I22"] = '=IF($C$5="","",SUMIFS(ST_ENCOURS,ST_LOT,$C$5))'
    cell(ws, "I22", None, bold=True, size=10, color=NAVY, fill=KPI_BG,
         border=BOX, align="right", fmt=NUM_KG0)

    # ---- blocs historiques ------------------------------------------------
    r = 24
    positions = {}
    for name, nrows, labels, rank in RECH_BLOCKS:
        section(ws, r, name, 11)
        hr = r + 1
        for j, lab in enumerate(labels):
            L = get_column_letter(2 + j)
            if lab:
                cell(ws, "%s%d" % (L, hr), lab, bold=True, size=8, color=WHITE,
                     fill=MUTED, align="center", wrap=True, border=BOX)
        ws.row_dimensions[hr].height = 26
        for i in range(nrows):
            rr = hr + 1 + i
            for j, lab in enumerate(labels):
                if not lab:
                    continue
                L = get_column_letter(2 + j)
                nm = SRC[rank][j]
                fmt = NUM_DATE if nm in DATES else (
                    NUM_KG0 if nm in KGS else None)
                cell(ws, "%s%d" % (L, rr), None, size=9, fill=WHITE,
                     border=BOX, fmt=fmt)
                if nm:
                    ws["%s%d" % (L, rr)] = (
                        '=IFERROR(INDEX({n},MATCH(ROW()-{o},{k},0)),"")'
                        .format(n=nm, o=hr, k=rank))
        positions[rank] = (hr, hr + nrows)
        r = hr + nrows + 2

    h0, h1 = positions["RES_RANG"]
    for rr in range(h0 + 1, h1 + 1):
        ws["J%d" % rr] = ('=IF($I%d="","",IFERROR(INDEX(EMPL_CLASSE,'
                          'MATCH($I%d,EMPL_KEY,0)),""))' % (rr, rr))
    h0, h1 = positions["RES_RANGLIEN"]
    for rr in range(h0 + 1, h1 + 1):
        ws["C%d" % rr] = ('=IF($B%d="","",IFERROR(INDEX(LOT_PRODE,'
                          'MATCH($B%d,LOT_KEY,0)),""))' % (rr, rr))
        ws["D%d" % rr] = ('=IF($B%d="","",IF(IFERROR(INDEX(LOT_ETAT,'
                          'MATCH($B%d,LOT_KEY,0)),"")="","",'
                          'INDEX(LOT_ETAT,MATCH($B%d,LOT_KEY,0))))'
                          % (rr, rr, rr))
        ws["E%d" % rr] = ('=IF($B%d="","",SUMIFS(MVT_NET,MVT_LOT,$B%d))'
                          % (rr, rr))
        ws["F%d" % rr] = ('=IF($B%d="","",IF($B%d=%s,'
                          '"LOT SOURCE / PARENT","ISSU DE SOUS-TRAITANCE"))'
                          % (rr, rr, PARENT_CELL))
        cell(ws, "E%d" % rr, None, size=9, fill=WHITE, border=BOX,
             fmt=NUM_KG0, align="right")
        for L in ("C", "D", "F"):
            cell(ws, "%s%d" % (L, rr), None, size=9, fill=WHITE, border=BOX)

    for rank, col in (("MVT_RANG", "J"), ("ST_RANG", "K"),
                      ("RES_RANG", "K"), ("QC_RANG", "J")):
        h0, h1 = positions[rank]
        cf_status(ws, "%s%d:%s%d" % (col, h0 + 1, col, h1))
    ws.freeze_panes = "A6"
    return ws
# ============================================================================
# ACCUEIL
# ============================================================================
NAV = [("LOTS", "1. Fiches lot"), ("OPERATIONS", "2. Operations de stock"),
       ("SOUS_TRAITANCE", "3. Envois sous-traitance"),
       ("ST_RESULTATS", "4. Resultats sous-traitance"),
       ("QUALITE", "5. Controles qualite"), ("STOCK", "6. Synthese stock"),
       ("RECHERCHE", "7. Recherche lot"), ("MOUVEMENTS", "8. Grand livre"),
       ("PARAM", "9. Parametres"), ("GUIDE", "10. Mode d'emploi")]

KPIS = [("STOCK INTERNE (kg)", "=STOCK!$D$4", NUM_KG0),
        ("STOCK EXTERNE (kg)", "=STOCK!$D$5", NUM_KG0),
        ("STOCK TOTAL (kg)", "=STOCK!$D$6", NUM_KG0),
        ("EN COURS SOUS-TRAITANCE (kg)", "=STOCK!$D$8", NUM_KG0),
        ("STOCK BLOQUE QUALITE (kg)", "=STOCK!$D$9", NUM_KG0),
        ("NOMBRE DE LOTS", "=COUNTA(LOT_KEY)", NUM_INT)]

CTRL = [
    ("Operations de stock en anomalie",
     '=COUNTIF(OPS_STATUT,"DONNEES MANQUANTES")'
     '+COUNTIF(OPS_STATUT,"STOCK INSUFFISANT")'
     '+COUNTIF(OPS_STATUT,"EMPLACEMENT INCOHERENT")', True),
    ("Envois sous-traitance bloquants",
     '=COUNTIF(ST_STATUT,"DONNEES MANQUANTES")'
     '+COUNTIF(ST_STATUT,"STOCK INSUFFISANT")'
     '+COUNTIF(ST_STATUT,"EMPLACEMENT INCOHERENT")'
     '+COUNTIF(ST_STATUT,"BLOQUE")', True),
    ("Resultats de sous-traitance en anomalie",
     '=COUNTIF(RES_STATUT,"DONNEES MANQUANTES")'
     '+COUNTIF(RES_STATUT,"BLOQUE")', True),
    ("Fiches lot en anomalie (doublon / incoherence)",
     '=COUNTIF(LOT_STATUT,"INCOHERENCE")'
     '+COUNTIF(LOT_STATUT,"LOT EN DOUBLE")', True),
    ("Lots avec stock negatif", '=COUNTIF(STOCK!$N$14:$N$%d,"<0")' % (13 + N_LOTS),
     True),
    ("Ecart matrice / grand livre (kg)", "=STOCK!$D$10", True),
    ("Reconciliation du grand livre",
     '=IF(STOCK!$H$%d="BILAN OK",0,1)' % (SYN + 7), True),
    ("Envois EN COURS (normal, en attente de resultat)",
     '=COUNTIF(ST_STATUT,"EN COURS")', False),
    ("Ecarts de sous-traitance A JUSTIFIER",
     '=COUNTIF(ST_STATUT,"A JUSTIFIER")', False),
    ("Controles qualite a verifier",
     '=COUNTIF(QC_STATUT,"DONNEES MANQUANTES")'
     '+COUNTIF(QC_STATUT,"EMPLACEMENT INCOHERENT")', False),
    ("Alertes qualite (histamine / temperature)",
     '=COUNTIF(QC_ALERTE,"?*")', False),
    ("Lignes de mouvement rejetees (non comptees en stock)",
     '=SUMPRODUCT((MVT_BRUT<>0)*(MVT_NET=0))', False),
]

STEPS = [
    ("1", "Creer le lot", "Onglet LOTS : une ligne par lot, une seule fois. "
     "C'est la seule saisie de l'identite (produit, fournisseur, origine, "
     "camion, etat, conservation)."),
    ("2", "Enregistrer les mouvements", "Onglet OPERATIONS : RECEPTION, "
     "TRANSFERT ou CONSOMMATION. On choisit le lot, les emplacements et la "
     "quantite - rien d'autre."),
    ("3", "Envoyer en sous-traitance", "Onglet SOUS_TRAITANCE. SOURCE = "
     "FOURNISSEUR pour une livraison directe (aucun stock consomme, aucune "
     "reception a saisir) ; SOURCE = STOCK EXISTANT pour un prelevement."),
    ("4", "Completer les resultats", "Onglet ST_RESULTATS, meme des jours plus "
     "tard. Un envoi sans resultat reste EN COURS : ce n'est pas une erreur."),
    ("5", "Justifier les ecarts", "QTE ENVOYEE - TOTAL SORTIE = ECART. Tant "
     "qu'un ecart positif n'est pas justifie, l'envoi reste A JUSTIFIER."),
    ("6", "Controler la qualite", "Onglet QUALITE : uniquement les mesures "
     "reellement faites. Un controle ne modifie jamais le stock."),
    ("7", "Corriger un inventaire", "Onglet OPERATIONS, type AJUSTEMENT : "
     "quantite signee sur l'emplacement de destination (negative = perte) "
     "et MOTIF obligatoire. C'est le seul moyen de corriger un stock."),
    ("8", "Consulter", "Onglet STOCK pour la synthese (par lot, par espece, "
     "par etat matiere, reconciliation), RECHERCHE pour tout savoir d'un lot, "
     "MOUVEMENTS pour l'historique complet."),
]

LEGEND = [(INPUT_BG, "Cellule a saisir"),
          (CALC_BG, "Cellule calculee - ne rien y ecrire"),
          (G_OK[0], "OK - la ligne est valide et comptee en stock"),
          (G_INFO[0], "EN COURS - etape normale, en attente de la suite"),
          (G_WARN[0], "A JUSTIFIER - il manque une explication"),
          (G_BAD[0], "ANOMALIE - la ligne n'est PAS comptee en stock")]


def build_accueil(wb):
    ws = wb.create_sheet("ACCUEIL", 0)
    title_bar(ws, "OCEAMIC  -  QUALITE, STOCK & SOUS-TRAITANCE",
              "Classeur unique - sans macro. Le stock est integralement "
              "reconstruit a partir de l'onglet MOUVEMENTS.", "I")
    for L, w in {"A": 2, "B": 34, "C": 15, "D": 3, "E": 34, "F": 15,
                 "G": 3, "H": 34, "I": 15}.items():
        ws.column_dimensions[L].width = w

    section(ws, 4, "NAVIGATION", 9)
    for i, (sheet, label) in enumerate(NAV):
        r = 5 + i // 3
        c = 2 + (i % 3) * 3
        L = get_column_letter(c)
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + 1)
        ws["%s%d" % (L, r)] = ('=HYPERLINK("#\'%s\'!A1","%s")'
                               % (sheet, label))
        cell(ws, "%s%d" % (L, r), None, bold=True, size=10, color="1B4F8F",
             fill=KPI_BG, border=BOX, align="left", indent=1)
        ws.row_dimensions[r].height = 20

    section(ws, 10, "TABLEAU DE BORD", 9)
    for i, (lab, f, fmt) in enumerate(KPIS):
        r = 11 + i // 3
        c = 2 + (i % 3) * 3
        L = get_column_letter(c)
        cell(ws, "%s%d" % (L, r), lab, bold=True, size=9, color=MUTED,
             fill=KPI_BG, border=BOX, align="left", indent=1)
        cell(ws, "%s%d" % (get_column_letter(c + 1), r), f, bold=True, size=12,
             color=NAVY, fill=KPI_BG, border=BOX, align="right", fmt=fmt)
        ws.row_dimensions[r].height = 22

    section(ws, 14, "CONTROLE D'INTEGRITE", 9)
    ws.merge_cells("B15:C15")
    ws["B15"] = ('=IF(SUM($C$17:$C$23)=0,"INTEGRITE DES DONNEES : OK",'
                 '"INTEGRITE DES DONNEES : A CORRIGER")')
    cell(ws, "B15", None, bold=True, size=12, color=WHITE, fill=NAVY,
         align="center", border=BOX)
    ws.row_dimensions[15].height = 24
    ws.conditional_formatting.add(
        "B15", FormulaRule(formula=['SUM($C$17:$C$23)>0'],
                           fill=PatternFill("solid", fgColor=G_BAD[1]),
                           font=Font(name=FONT, size=12, bold=True,
                                     color=WHITE)))
    ws.merge_cells("E15:I15")
    cell(ws, "E15", "Les 6 premieres lignes ci-dessous sont bloquantes : tant "
         "qu'elles ne sont pas a zero, le stock affiche n'est pas fiable a "
         "100%. Les suivantes sont informatives.", size=8, color=MUTED,
         align="left", wrap=True, indent=1)

    for i, (lab, f, blocking) in enumerate(CTRL):
        r = 17 + i
        cell(ws, "B%d" % r, lab, size=9, bold=blocking,
             color=INK if blocking else MUTED,
             fill=WHITE, border=BOX, align="left", indent=1)
        cell(ws, "C%d" % r, f, bold=True, size=10, color=INK, fill=WHITE,
             border=BOX, align="center", fmt='#,##0.###;-#,##0.###;"-"')
        ws.conditional_formatting.add(
            "C%d" % r,
            CellIsRule(operator="notEqual", formula=["0"],
                       fill=PatternFill("solid",
                                        fgColor=G_BAD[0] if blocking
                                        else G_WARN[0]),
                       font=Font(name=FONT, size=10, bold=True,
                                 color=G_BAD[1] if blocking else G_WARN[1])))

    ws.merge_cells("E17:I17")
    cell(ws, "E17", "LEGENDE DES COULEURS", bold=True, size=10, color=NAVY,
         fill=TEAL_SOFT, align="left", indent=1)
    for i, (bg, txt) in enumerate(LEGEND):
        r = 18 + i
        cell(ws, "E%d" % r, "", fill=bg, border=BOX)
        ws.merge_cells("F%d:I%d" % (r, r))
        cell(ws, "F%d" % r, txt, size=9, color=INK, align="left", indent=1)

    section(ws, 29, "MODE OPERATOIRE EN 7 ETAPES", 9)
    for i, (n, t, d) in enumerate(STEPS):
        r = 30 + i
        cell(ws, "B%d" % r, "%s. %s" % (n, t), bold=True, size=9, color=NAVY,
             fill=KPI_BG, border=BOX, align="left", indent=1)
        ws.merge_cells("C%d:I%d" % (r, r))
        cell(ws, "C%d" % r, d, size=9, color=INK, fill=WHITE, border=BOX,
             align="left", wrap=True, indent=1)
        ws.row_dimensions[r].height = 28
    ws.sheet_view.showGridLines = False
    return ws


# ============================================================================
# GUIDE
# ============================================================================
GUIDE_TXT = [
    ("T", "REGLES DE GESTION APPLIQUEES PAR LE CLASSEUR"),
    ("S", "1. Identite du lot - aucune double saisie"),
    ("P", "L'onglet LOTS est l'unique endroit ou l'on decrit un lot. Partout "
          "ailleurs on ne choisit que le CODE du lot : produit, fournisseur, "
          "origine, etat et conservation sont repris automatiquement."),
    ("P", "Si deux fiches portent le meme code lot avec un produit ou un "
          "fournisseur different, la colonne STATUT FICHE affiche "
          "INCOHERENCE. Si les deux fiches sont identiques, elle affiche "
          "LOT EN DOUBLE. Dans les deux cas la ligne apparait en rouge."),
    ("P", "Le LOT INTERNE est la cle du stock : obligatoire, unique et "
          "definitif. Le LOT EXTERNE (numero du fournisseur) est un attribut "
          "cherchable et tracable ; il peut etre partage par plusieurs lots "
          "internes. Si aucun numero interne n'est encore attribue, reprenez "
          "le numero externe comme LOT INTERNE et ne le changez plus."),
    ("P", "QUALITE et MOULE acceptent la saisie libre : les listes "
          "deroulantes ne sont qu'une aide. Les formats de repartition "
          "(A=20 B=40 C=40, 20/24=14 26/30=55 38/50=31) passent tels quels."),
    ("S", "2. Classes de stock"),
    ("P", "Une seule dimension : INTERNE ou EXTERNE. OCEAMIC 2 est INTERNE ; "
          "SARMA, DAMSA, COFRIGOP, COFRIGOB et tout autre entrepot ou usine "
          "tiers sont EXTERNE. CONSOMMATION n'est PAS une classe de stock : "
          "c'est une operation."),
    ("S", "3. Operations physiques"),
    ("P", "RECEPTION : ajoute la quantite sur l'emplacement de destination."),
    ("P", "TRANSFERT : retire la quantite d'un emplacement et ajoute la MEME "
          "quantite sur un autre. Une seule ligne de saisie genere deux "
          "lignes de mouvement."),
    ("P", "CONSOMMATION : retire uniquement la quantite saisie. 1 000 kg en "
          "stock, 300 kg consommes, il reste 700 kg. Le classeur ne solde "
          "jamais un lot automatiquement."),
    ("P", "AJUSTEMENT : correction d'inventaire sur l'emplacement de "
          "destination. La quantite est signee (negative = perte, positive = "
          "gain) et le MOTIF est obligatoire. Un ajustement negatif ne peut "
          "pas depasser le disponible."),
    ("S", "4. Moteur de stock"),
    ("P", "Toute operation valide alimente l'onglet MOUVEMENTS, qui est le "
          "seul historique. Le stock n'est jamais stocke : il est recalcule "
          "par ENTREES - SORTIES sur ce grand livre."),
    ("P", "MOUVEMENTS comporte deux colonnes de quantite. QTE DECLAREE sert a "
          "calculer le disponible avant operation. QTE EN STOCK vaut zero si "
          "la ligne est en anomalie : une ligne invalide n'affecte donc "
          "jamais le stock."),
    ("P", "Consequence directe : le stock ne peut pas devenir negatif. Si une "
          "sortie depasse le disponible, la ligne passe en STOCK INSUFFISANT, "
          "elle est ecartee du stock et signalee en rouge sur la ligne, dans "
          "le grand livre et sur l'ACCUEIL. Sans macro, Excel ne peut pas "
          "empecher physiquement la frappe : il la detecte et la neutralise."),
    ("P", "Les doublons de mouvement sont impossibles a creer a la main (le "
          "grand livre est entierement calcule) ; une saisie identique en "
          "double dans OPERATIONS est signalee par la colonne CONTROLE."),
    ("S", "5. Sous-traitance"),
    ("P", "SOURCE = FOURNISSEUR : la marchandise part du fournisseur "
          "directement chez le sous-traitant. Ni emplacement source, ni "
          "reception, ni stock consomme. L'enregistrement de la "
          "sous-traitance suffit."),
    ("P", "SOURCE = STOCK EXISTANT : emplacement source et quantite envoyee "
          "obligatoires ; le disponible est verifie automatiquement et seule "
          "la quantite envoyee est retiree."),
    ("P", "Saisie en deux temps : on enregistre l'envoi, le statut passe a "
          "EN COURS. Ce n'est pas une erreur. Les resultats sont saisis plus "
          "tard dans ST_RESULTATS."),
    ("P", "Un envoi peut produire plusieurs resultats, avec des etats, des "
          "qualites et des moules differents, et des destinations "
          "differentes (retour OCEAMIC 2 ou entrepot externe)."),
    ("P", "ECART A JUSTIFIER = QTE ENVOYEE - TOTAL SORTIE. Au-dela de la "
          "tolerance definie dans PARAM, l'envoi reste A JUSTIFIER tant que "
          "la colonne JUSTIFICATION ECART est vide. Aucune quantite ne "
          "disparait en silence."),
    ("P", "Qualite, moule et histamine sont acceptes a l'entree comme a la "
          "sortie, mais ne sont jamais obligatoires."),
    ("S", "6. Controles qualite"),
    ("P", "Un controle qualite ne modifie JAMAIS une quantite. On saisit "
          "seulement ce qui a ete mesure. Le resultat appartient au lot "
          "physiquement controle ; un lot lie reste tracable mais n'herite "
          "pas du resultat."),
    ("P", "Si le lot n'est pas present a l'emplacement indique, le statut "
          "affiche EMPLACEMENT INCOHERENT : c'est un avertissement, le "
          "controle est conserve et le stock reste inchange."),
    ("S", "7. Statuts"),
    ("P", "OK : ligne valide, comptee en stock."),
    ("P", "EN COURS : envoi enregistre, resultats non connus."),
    ("P", "DONNEES MANQUANTES : champs indispensables absents ; ligne non "
          "comptee."),
    ("P", "STOCK INSUFFISANT : sortie superieure au disponible ; ligne non "
          "comptee."),
    ("P", "EMPLACEMENT INCOHERENT : le lot n'existe pas a cet emplacement."),
    ("P", "A JUSTIFIER : ecart de sous-traitance sans justification."),
    ("P", "BLOQUE : total des sorties superieur a la quantite envoyee, ou "
          "envoi parent en anomalie."),
    ("S", "8. Capacites et limites"),
    ("P", "LOTS : %d lignes. OPERATIONS : %d. SOUS_TRAITANCE : %d. "
          "ST_RESULTATS : %d. QUALITE : %d. MOUVEMENTS : %d lignes generees. "
          "Matrice STOCK : %d emplacements."
          % (N_LOTS, N_OPS, N_ST, N_RES, N_QC, MV_END - MV_B1 + 1,
             N_COLS_EMPL)),
    ("P", "Pour agrandir : copier la derniere ligne d'un onglet de saisie "
          "vers le bas, puis etendre de la meme facon les blocs "
          "correspondants de MOUVEMENTS et les plages nommees."),
    ("P", "Aucune macro, aucun bouton, aucune protection de cellule : le "
          "classeur fonctionne sur PC, sur Excel Online et sur mobile."),
    ("S", "9. Ordre de saisie recommande"),
    ("P", "LOTS -> OPERATIONS -> SOUS_TRAITANCE -> ST_RESULTATS -> QUALITE. "
          "Consultation : STOCK, RECHERCHE, MOUVEMENTS."),
]


def build_guide(wb):
    ws = wb.create_sheet("GUIDE")
    title_bar(ws, "MODE D'EMPLOI ET REGLES DE GESTION",
              "A lire une fois. Tout le fonctionnement du classeur tient dans "
              "cette page.", "F")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 120
    r = 4
    for kind, txt in GUIDE_TXT:
        if kind == "T":
            section(ws, r, txt, 6, size=11)
            r += 2
        elif kind == "S":
            cell(ws, "B%d" % r, txt, bold=True, size=10, color=NAVY,
                 align="left")
            ws.row_dimensions[r].height = 22
            r += 1
        else:
            cell(ws, "B%d" % r, txt, size=9, color=INK, align="left",
                 wrap=True)
            ws.row_dimensions[r].height = max(16, 13 * (len(txt) // 105 + 1))
            r += 1
    ws.sheet_view.showGridLines = False
    return ws
D = dt.date

# ---------------------------------------------------------------- jeux d'essai
DEMO = {
    "LOTS": [
        ["255", "L-EXT-255", "Sardine", "Fournisseur local", "Laayoune",
         "12345-A-6", "ENTIER", "CONGELE", "A", "20/24", D(2026, 8, 3), None,
         "Lot de reference - scenarios 1, 5, 8, 11, 12, 14, 15"],
        ["256", "L-EXT-256", "Maquereau", "Producteur Boujdour", "Boujdour",
         "22876-B-1", "ENTIER", "CONGELE", "B", "26/30", D(2026, 8, 4), None,
         "Scenarios 2 et 4 - stock externe puis transfert"],
        ["257", "L-EXT-255", "Sardine", "Fournisseur local", "Laayoune",
         "33111-C-2", "ENTIER", "FRAIS", "A", "20/24", D(2026, 8, 5), None,
         "Scenario 3 - meme lot externe que 255 (cas reel a deux lots)"],
        ["258", "L-EXT-258", "Sardine", "Fournisseur local", "tan tan",
         "44900-E-7", "ENTIER", "CONGELE", "B", "20/24", D(2026, 8, 6), None,
         "Scenario 13 - tentative de sortie superieure au stock"],
        ["300", "L-EXT-300", "Anchois", "Producteur Dakhla", "Dakhla",
         "44222-D-3", "ENTIER", "CONGELE", "A", "20/24", D(2026, 8, 7), None,
         "Scenario 6 - livraison directe fournisseur vers sous-traitant"],
        ["301", "L-EXT-301", "Maquereau", "Producteur Boujdour", "Boujdour",
         "55333-F-8", "ENTIER", "CONGELE", "A", "26/30", D(2026, 8, 10), None,
         "Scenario 7 - sous-traitance EN COURS, puis blocage qualite"],
        ["302", "L-EXT-302", "Chinchard", "Producteur Dakhla", "Dakhla",
         "66444-G-9", "ENTIER", "CONGELE", "A", "26/30", D(2026, 8, 17), None,
         "Scenario 21 - ajustement d'inventaire"],
        ["255-HG", None, None, None, None, None, "HG", "CONGELE", "A",
         "20/24", D(2026, 8, 14), "255",
         "Issu de ST-001 - identite heritee du lot 255"],
        ["255-HGT", None, None, None, None, None, "HGT", "CONGELE", "B",
         "26/30", D(2026, 8, 14), "255",
         "Issu de ST-001 - envoye en stock externe COFRIGOP"],
        ["300-FIL", None, None, None, None, None, "FILET", "CONGELE", "A",
         "20/24", D(2026, 8, 13), "300", "Issu de ST-002 - retour OCEAMIC 2"],
    ],
    "OPERATIONS": [
        [D(2026, 8, 3), "RECEPTION", "255", None, "OCEAMIC 2", 1000,
         "Reception usine", "Scenario 1 - nouveau lot interne"],
        [D(2026, 8, 4), "RECEPTION", "256", None, "SARMA", 2000,
         "Reception entrepot externe", "Scenario 2 - lot externe"],
        [D(2026, 8, 5), "RECEPTION", "257", None, "OCEAMIC 2", 800,
         "Reception usine", None],
        [D(2026, 8, 6), "RECEPTION", "258", None, "OCEAMIC 2", 200,
         "Reception usine", None],
        [D(2026, 8, 8), "CONSOMMATION", "257", "OCEAMIC 2", None, 300,
         "Production conserve", "Scenario 3 - il doit rester 500 kg"],
        [D(2026, 8, 9), "TRANSFERT", "256", "SARMA", "OCEAMIC 2", 800,
         "Rapatriement usine", "Scenario 4 - transfert partiel"],
        [D(2026, 8, 10), "RECEPTION", "301", None, "OCEAMIC 2", 400,
         "Reception usine", None],
        [D(2026, 8, 11), "CONSOMMATION", "258", "OCEAMIC 2", None, 500,
         "Production conserve",
         "Scenario 13 - doit etre refuse : stock disponible 200 kg"],
        [D(2026, 8, 17), "RECEPTION", "302", None, "OCEAMIC 2", 1000,
         "Reception usine", None],
        [D(2026, 8, 18), "AJUSTEMENT", "302", None, "OCEAMIC 2", -25,
         "Ecart d'inventaire constate au comptage",
         "Scenario 21 - ajustement negatif, il doit rester 975 kg"],
    ],
    "SOUS_TRAITANCE": [
        [D(2026, 8, 12), "USINE SOUS-TRAITANTE", "STOCK EXISTANT", "255",
         "OCEAMIC 2", 500, "A=60 B=40", "20/24", 45, D(2026, 8, 14), None,
         None, "Scenarios 5, 8, 9, 10, 11"],
        [D(2026, 8, 7), "USINE SOUS-TRAITANTE", "FOURNISSEUR", "300", None,
         600, "A", "20/24", None, D(2026, 8, 13),
         "Perte de process au filetage, pesee et controlee", "PERTE PROCESS",
         "Scenario 6 - aucun stock OCEAMIC consomme"],
        [D(2026, 8, 15), "CONGELATION", "STOCK EXISTANT", "301",
         "OCEAMIC 2", 200, None, None, None, None, None, None,
         "Scenario 7 - reste EN COURS tant qu'aucun resultat n'est saisi"],
    ],
    "ST_RESULTATS": [
        ["ST-001", D(2026, 8, 14), "HG", "255-HG", 300, "A", "20/24", 40,
         "OCEAMIC 2", "Scenarios 8 et 9 - retour OCEAMIC 2"],
        ["ST-001", D(2026, 8, 14), "HGT", "255-HGT", 150, "B", "26/30", None,
         "COFRIGOP", "Scenarios 8 et 10 - resultat vers stock externe"],
        ["ST-002", D(2026, 8, 13), "FILET", "300-FIL", 590, "A", "20/24",
         None, "OCEAMIC 2", "Scenario 6 - ecart de 10 kg justifie"],
    ],
    "QUALITE": [
        [D(2026, 8, 13), "255", "OCEAMIC 2", "CONTROLE MOUVEMENT", -19, 45,
         "A=20 B=40 C=40", "20/24=14 26/30=55 38/50=31", "Aucun defaut",
         "ACCEPTER", "K. BENALI",
         "Scenario 12 - aucun impact sur le stock"],
        [D(2026, 8, 14), "257", "COFRIGOB", "STOCK", 2, None, "A", None,
         None, "ACCEPTER", "K. BENALI",
         "Emplacement incoherent : le lot 257 est a OCEAMIC 2"],
        [D(2026, 8, 16), "256", "OCEAMIC 2", "RECEPTION", -18, 180,
         "A=10 B=60 C=30", "26/30=70 38/50=30",
         "Aspect legerement decolore", "A RECONTROLER", "K. BENALI",
         "Alerte histamine au-dessus du seuil"],
        [D(2026, 8, 19), "301", "OCEAMIC 2", "STOCK", -17, 210, None, None,
         "Suspicion histamine", "BLOQUER", "K. BENALI",
         "Scenario 22 - lot bloque qualite : signale, stock inchange"],
    ],
}

EX = "LIGNE EXEMPLE - a supprimer avant la mise en service"
SAMPLE = {
    "LOTS": [["EXEMPLE-001", "L-EXT-001", "Sardine", "Fournisseur local",
              "Laayoune", "12345-A-6", "ENTIER", "CONGELE", "A", "20/24",
              D(2026, 9, 1), None, EX]],
    "OPERATIONS": [[D(2026, 9, 1), "RECEPTION", "EXEMPLE-001", None,
                    "OCEAMIC 2", 1000, "Reception usine", EX]],
    "SOUS_TRAITANCE": [[D(2026, 9, 1), "USINE SOUS-TRAITANTE",
                        "STOCK EXISTANT", "EXEMPLE-001", "OCEAMIC 2", 200,
                        "A", "20/24", 45, D(2026, 9, 2),
                        "Perte de process pesee et controlee",
                        "PERTE PROCESS", EX]],
    "ST_RESULTATS": [["ST-001", D(2026, 9, 2), "HG", None, 190, "A", "20/24",
                      None, "OCEAMIC 2",
                      EX + " - LOT RESULTAT vide = on garde le lot source"]],
    "QUALITE": [[D(2026, 9, 1), "EXEMPLE-001", "OCEAMIC 2",
                 "CONTROLE MOUVEMENT", -19, 45, "A=30 B=50 C=20",
                 "20/24=40 26/30=60", "Aucun defaut", "ACCEPTER", "QUALITE",
                 EX]],
}

SEED_START = {"LOTS": "A", "OPERATIONS": "B", "SOUS_TRAITANCE": "B",
              "ST_RESULTATS": "B", "QUALITE": "B"}


def seed(wb, data):
    from openpyxl.utils import column_index_from_string
    for sheet, rows in data.items():
        ws = wb[sheet]
        c0 = column_index_from_string(SEED_START[sheet])
        for i, row in enumerate(rows):
            for j, v in enumerate(row):
                if v is not None:
                    ws.cell(row=FIRST + i, column=c0 + j).value = v


# ============================================================================
# TESTS  (uniquement dans la version de recette)
# ============================================================================
TESTS = [
    ("1", "Reception d'un nouveau lot interne : 255, 1 000 kg sur OCEAMIC 2",
     1000, '=SUMIFS(MVT_NET,MVT_LOT,"255",MVT_EMPL,"OCEAMIC 2",'
           'MVT_OP,"RECEPTION")'),
    ("2", "Reception d'un lot externe : 256, 2 000 kg sur SARMA (EXTERNE)",
     2000, '=SUMIFS(MVT_NET,MVT_LOT,"256",MVT_EMPL,"SARMA",'
           'MVT_OP,"RECEPTION")'),
    ("3", "Consommation partielle : 800 kg - 300 kg -> il reste 500 kg",
     500, '=SUMIFS(MVT_NET,MVT_LOT,"257",MVT_EMPL,"OCEAMIC 2")'),
    ("4a", "Transfert partiel : SARMA passe de 2 000 a 1 200 kg",
     1200, '=SUMIFS(MVT_NET,MVT_LOT,"256",MVT_EMPL,"SARMA")'),
    ("4b", "Transfert partiel : OCEAMIC 2 recoit les 800 kg",
     800, '=SUMIFS(MVT_NET,MVT_LOT,"256",MVT_EMPL,"OCEAMIC 2")'),
    ("5", "Sous-traitance depuis stock : 1 000 - 500 envoyes -> 500 kg",
     500, '=SUMIFS(MVT_NET,MVT_LOT,"255",MVT_EMPL,"OCEAMIC 2")'),
    ("6a", "Fournisseur direct : aucun stock OCEAMIC consomme sur le lot 300",
     0, '=SUMIFS(MVT_NET,MVT_LOT,"300")'),
    ("6b", "Fournisseur direct : aucune ligne de sortie generee pour ST-002",
     0, '=ABS(SUMIFS(MVT_BRUT,MVT_REF,"ST-002",MVT_SENS,"SORTIE"))'),
    ("7", "Envoi sans resultat : ST-003 reste EN COURS",
     "EN COURS", '=INDEX(ST_STATUT,MATCH("ST-003",ST_ID,0))'),
    ("8a", "Un envoi, plusieurs resultats : ST-001 a 2 resultats",
     2, '=INDEX(ST_NBRES,MATCH("ST-001",ST_ID,0))'),
    ("8b", "Un envoi, plusieurs resultats : total sorti 300 + 150 = 450 kg",
     450, '=INDEX(ST_TOTSORTIE,MATCH("ST-001",ST_ID,0))'),
    ("9", "Retour OCEAMIC 2 : lot 255-HG, 300 kg en stock interne",
     300, '=SUMIFS(MVT_NET,MVT_LOT,"255-HG",MVT_EMPL,"OCEAMIC 2")'),
    ("10", "Resultat vers stock externe : lot 255-HGT, 150 kg sur COFRIGOP",
     150, '=SUMIFS(MVT_NET,MVT_LOT,"255-HGT",MVT_EMPL,"COFRIGOP")'),
    ("11a", "Ecart a justifier : 500 - 450 = 50 kg",
     50, '=INDEX(ST_ECART,MATCH("ST-001",ST_ID,0))'),
    ("11b", "Ecart non justifie : ST-001 au statut A JUSTIFIER",
     "A JUSTIFIER", '=INDEX(ST_STATUT,MATCH("ST-001",ST_ID,0))'),
    ("11c", "Ecart justifie : ST-002 (10 kg justifies) au statut OK",
     "OK", '=INDEX(ST_STATUT,MATCH("ST-002",ST_ID,0))'),
    ("12a", "Controle qualite : aucune ligne de mouvement generee",
     0, '=COUNTIF(MVT_SRC,"QUALITE")'),
    ("12b", "Controle qualite : le stock du lot 255 reste inchange",
     500, '=SUMIFS(MVT_NET,MVT_LOT,"255",MVT_EMPL,"OCEAMIC 2")'),
    ("12c", "Controle a un emplacement ou le lot est absent -> avertissement",
     "EMPLACEMENT INCOHERENT", '=INDEX(QC_STATUT,MATCH("257",QC_LOT,0))'),
    ("13a", "Sortie superieure au stock : ligne refusee (STOCK INSUFFISANT)",
     "STOCK INSUFFISANT", '=OPERATIONS!$Q$12'),
    ("13b", "Sortie refusee : le stock du lot 258 reste a 200 kg",
     200, '=SUMIFS(MVT_NET,MVT_LOT,"258")'),
    ("14a", "Lot reutilise : espece reprise automatiquement, sans re-saisie",
     "Sardine", '=OPERATIONS!$K$12'),
    ("14b", "Lot reutilise : lot externe repris automatiquement",
     "L-EXT-302", '=OPERATIONS!$J$13'),
    ("15a", "Tracabilite : 2 lots lies au lot 255 (sans doublon)",
     2, '=COUNT(RES_RANGLIEN)'),
    ("15b", "Tracabilite : parent du lot 255-HG retrouve automatiquement",
     "255", '=IFERROR(INDEX(RES_LOTSRC,MATCH("255-HG",RES_LOTEFF,0)),"")'),
    ("15c", "Lot externe partage : 3 autres lots internes sous L-EXT-255",
     3, '=COUNT(LOT_RANGEXT)'),
    ("16", "Integrite : aucun stock negatif dans la matrice",
     0, '=COUNTIF(STOCK!$N$%d:$N$%d,"<0")' % (R0_M, R1_M)),
    ("17", "Integrite : matrice STOCK = grand livre MOUVEMENTS (ecart 0 kg)",
     0, '=STOCK!$D$10'),
    ("18", "Integrite : stock total interne attendu 4 065 kg",
     4065, '=SUMIFS(MVT_NET,MVT_CLASSE,"INTERNE")'),
    ("19", "Integrite : stock total externe attendu 1 350 kg",
     1350, '=SUMIFS(MVT_NET,MVT_CLASSE,"EXTERNE")'),
    ("20", "Alerte qualite : histamine 180 PPM signalee sur le lot 256",
     1, '=IF(LEN(INDEX(QC_ALERTE,MATCH("256",QC_LOT,0)))>0,1,0)'),
    ("21a", "Ajustement negatif : 1 000 - 25 -> il reste 975 kg",
     975, '=SUMIFS(MVT_NET,MVT_LOT,"302")'),
    ("21b", "Ajustement enregistre comme AJUSTEMENT dans le grand livre",
     -25, '=SUMIFS(MVT_NET,MVT_OP,"AJUSTEMENT")'),
    ("21c", "Ajustement valide (motif renseigne)", "OK", '=OPERATIONS!$Q$14'),
    ("22a", "Blocage qualite : lot 301 signale BLOQUER, 200 kg bloques",
     200, '=STOCK!$D$9'),
    ("22b", "Blocage qualite : le stock du lot 301 est inchange",
     200, '=SUMIFS(MVT_NET,MVT_LOT,"301")'),
    ("23a", "Synthese par espece : Sardine 1 650 kg",
     1650, '=SUMIFS(MVT_NET,MVT_PROD,"Sardine")'),
    ("23b", "Synthese par etat matiere : ENTIER 4 375 kg",
     4375, '=SUMIFS(MVT_NET,MVT_ETAT,"ENTIER")'),
    ("24", "Reconciliation : ENTREES - SORTIES +/- AJUSTEMENTS = STOCK ACTUEL",
     "BILAN OK", '=STOCK!$H$%d' % (SYN + 7)),
]

def build_tests(wb):
    ws = wb.create_sheet("TESTS")
    title_bar(ws, "RECETTE  -  SCENARIOS DE TEST",
              "Chaque ligne compare un resultat attendu au resultat reellement "
              "calcule par le classeur. Tout doit etre PASS.", "F")
    for L, w in {"A": 2, "B": 7, "C": 74, "D": 24, "E": 24, "F": 12}.items():
        ws.column_dimensions[L].width = w
    ws.merge_cells("B4:C4")
    ws["B4"] = ('=IF(COUNTIF($F$7:$F$%d,"FAIL")=0,"RESULTAT GLOBAL : '
                '%d / %d TESTS PASS","RESULTAT GLOBAL : "'
                '&COUNTIF($F$7:$F$%d,"FAIL")&" TEST(S) EN ECHEC")'
                % (6 + len(TESTS), len(TESTS), len(TESTS), 6 + len(TESTS)))
    cell(ws, "B4", None, bold=True, size=12, color=WHITE, fill=NAVY,
         align="center", border=BOX)
    ws.row_dimensions[4].height = 24
    ws.conditional_formatting.add(
        "B4", FormulaRule(formula=['COUNTIF($F$7:$F$%d,"FAIL")>0'
                                   % (6 + len(TESTS))],
                          fill=PatternFill("solid", fgColor=G_BAD[1]),
                          font=Font(name=FONT, size=12, bold=True,
                                    color=WHITE)))
    for L, lab in (("B", "N"), ("C", "SCENARIO TESTE"), ("D", "ATTENDU"),
                   ("E", "OBTENU"), ("F", "RESULTAT")):
        cell(ws, "%s6" % L, lab, bold=True, size=9, color=WHITE, fill=TEAL,
             align="center", border=BOX)
    for i, (n, desc, exp, formula) in enumerate(TESTS):
        r = 7 + i
        cell(ws, "B%d" % r, n, bold=True, size=9, color=NAVY, fill=WHITE,
             border=BOX, align="center")
        cell(ws, "C%d" % r, desc, size=9, color=INK, fill=WHITE, border=BOX,
             align="left", indent=1)
        cell(ws, "D%d" % r, exp, size=9, color=MUTED, fill=CALC_BG,
             border=BOX, align="center",
             fmt=NUM_KG0 if isinstance(exp, (int, float)) else None)
        cell(ws, "E%d" % r, formula, bold=True, size=9, color=INK, fill=WHITE,
             border=BOX, align="center",
             fmt=NUM_KG0 if isinstance(exp, (int, float)) else None)
        ws["F%d" % r] = ('=IF(ISNUMBER($D{r}),IF(ABS($E{r}-$D{r})<0.001,'
                         '"PASS","FAIL"),IF(EXACT(TRIM($E{r}),TRIM($D{r})),'
                         '"PASS","FAIL"))').format(r=r)
        cell(ws, "F%d" % r, None, bold=True, size=9, fill=WHITE, border=BOX,
             align="center")
    rng = "F7:F%d" % (6 + len(TESTS))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=['"PASS"'],
        fill=PatternFill("solid", fgColor=G_OK[0]),
        font=Font(name=FONT, size=9, bold=True, color=G_OK[1])))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=['"FAIL"'],
        fill=PatternFill("solid", fgColor=G_BAD[0]),
        font=Font(name=FONT, size=9, bold=True, color=G_BAD[1])))
    r = 8 + len(TESTS)
    ws.merge_cells("B%d:F%d" % (r, r))
    cell(ws, "B%d" % r,
         "Note : dans ce classeur de recette, l'ACCUEIL affiche "
         "\"INTEGRITE DES DONNEES : A CORRIGER\". C'est VOULU - le scenario "
         "13 laisse volontairement une consommation superieure au stock, "
         "refusee par le systeme. Le classeur de production, lui, demarre a "
         "OK.", size=9, color=MUTED, fill=KPI_BG, border=BOX, align="left",
         wrap=True, indent=1)
    ws.row_dimensions[r].height = 30
    ws.freeze_panes = "A7"
    return ws


# ============================================================================
def build(path, demo=False):
    wb = Workbook()
    wb.remove(wb.active)
    build_param(wb)
    build_lots(wb)
    build_operations(wb)
    build_st(wb)
    build_res(wb)
    build_qualite(wb)
    build_mouvements(wb)
    build_stock(wb)
    build_recherche(wb)
    build_guide(wb)
    build_accueil(wb)
    add_names(wb)
    seed(wb, DEMO if demo else SAMPLE)
    if demo:
        build_tests(wb)
        wb["RECHERCHE"]["C4"] = "255"
    else:
        wb["RECHERCHE"]["C4"] = "EXEMPLE-001"
    order = ["ACCUEIL", "LOTS", "OPERATIONS", "SOUS_TRAITANCE",
             "ST_RESULTATS", "QUALITE", "STOCK", "RECHERCHE", "MOUVEMENTS",
             "PARAM", "GUIDE"] + (["TESTS"] if demo else [])
    wb._sheets = [wb[n] for n in order]
    for name, last, nrows in (("LOTS", "U", N_LOTS), ("OPERATIONS", "R", N_OPS),
                              ("SOUS_TRAITANCE", "AB", N_ST),
                              ("ST_RESULTATS", "Z", N_RES),
                              ("QUALITE", "T", N_QC),
                              ("MOUVEMENTS", "N", MV_END - MV_B1 + 1)):
        ws = wb[name]
        top = HDR if name != "MOUVEMENTS" else HDR
        ws.auto_filter.ref = "A%d:%s%d" % (top, last, top + nrows)
    wb.calculation.fullCalcOnLoad = True
    for ws in wb.worksheets:
        ws.sheet_properties.tabColor = NAVY if ws.title in (
            "ACCUEIL", "STOCK", "RECHERCHE") else TEAL
    wb.save(path)
    return path


if __name__ == "__main__":
    import sys
    print(build(sys.argv[1], demo=(len(sys.argv) > 2)))
